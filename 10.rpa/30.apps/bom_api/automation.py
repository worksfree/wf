import os
import sys
import time
import traceback
from collections import Counter, OrderedDict
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

# automation.py
# Requirements: pywin32, openpyxl, pillow
# pip install pywin32 openpyxl pillow


import win32com.client

# --- User configuration ---
ASSEMBLY_FOLDER = r"D:\drive_files\10.worksfree\10.rpa\30.apps\bom_api\assemblies"  # change to folder with .sldasm files
OUTPUT_XLSX = r"D:\drive_files\10.worksfree\10.rpa\30.apps\bom_api\bom_with_thumbs.xlsx"
THUMBNAIL_FOLDER = r"D:\drive_files\10.worksfree\10.rpa\30.apps\bom_api\thumbnails"
SOLIDWORKS_VISIBLE = True
# ---------------------------

swDocPART = 1
swDocASSEMBLY = 2
swDocDRAWING = 3

os.makedirs(THUMBNAIL_FOLDER, exist_ok=True)


def start_solidworks():
    sw = win32com.client.Dispatch("SldWorks.Application")
    sw.Visible = SOLIDWORKS_VISIBLE
    return sw


def open_document(sw, path):
    # OpenDoc6(fileName, docType, options, configuration, errors, warnings)
    errors = (
        win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        if "pythoncom" in globals()
        else 0
    )
    warnings = 0
    try:
        # Try OpenDoc6; using 0 for options/configuration for simplicity
        model = sw.OpenDoc6(path, swDocASSEMBLY, 0, "", errors, warnings)
    except Exception:
        # fallback to OpenDoc (older)
        model = sw.OpenDoc(path, swDocASSEMBLY)
    return model


def save_thumbnail_from_model(model_doc, out_path):
    """
    Try to export a thumbnail image for a model document.
    Approach: try ModelDoc2::SaveAs to a PNG file (SolidWorks typically supports exporting images).
    If that fails, return False.
    """
    try:
        # Some SolidWorks setups accept SaveAs to image formats. Try PNG first.
        # SaveAs signature can be variable; try simple call and handle exceptions.
        try:
            # Many setups accept SaveAs(fileName)
            model_doc.SaveAs(out_path)
        except Exception:
            # Try SaveAs with more parameters if available (some bindings require extras)
            try:
                errors = 0
                warnings = 0
                model_doc.SaveAs(out_path, 0, 0, None, errors, warnings)
            except Exception:
                model_doc.SaveAs(out_path)
        # Verify file exists
        if os.path.isfile(out_path):
            # Ensure image can be opened (normalize format)
            try:
                img = Image.open(out_path)
                img.save(out_path)  # re-save to normalize if necessary
            except Exception:
                pass
            return True
    except Exception:
        pass
    return False


def gather_bom_from_assembly(assembly_model):
    """
    Returns OrderedDict keyed by component path (or name) with dict {name, path, qty, model_doc}
    """
    bom = OrderedDict()
    try:
        # Attempt to get the AssemblyDoc interface and components
        # Many SolidWorks COM wrappers expose GetComponents(includeSubcomponents)
        components = None
        try:
            components = assembly_model.GetComponents(True)
        except Exception:
            try:
                assy = assembly_model  # fallback
                components = assy.GetComponents(False)
            except Exception:
                components = None

        if not components:
            return bom

        # components is a tuple/list of Component2 objects
        paths = []
        comp_objs = []
        for comp in components:
            try:
                comp_path = comp.GetPathName() or comp.GetModelName() or comp.Name2
            except Exception:
                comp_path = getattr(comp, "Name2", None) or None
            comp_name = getattr(comp, "Name2", None) or os.path.basename(comp_path or "unknown")
            paths.append(comp_path or comp_name)
            comp_objs.append((comp_path or comp_name, comp_name, comp))
        counts = Counter(paths)
        # Build BOM entries in original order
        seen = set()
        for key, name, comp in comp_objs:
            if key in seen:
                continue
            seen.add(key)
            bom[key] = {
                "name": name,
                "path": key,
                "qty": counts.get(key, 1),
                "component": comp,
            }
    except Exception:
        traceback.print_exc()
    return bom


def write_bom_to_excel(bom_dict, assembly_name, workbook, thumbnail_folder):
    ws = workbook.create_sheet(title=assembly_name[:31])
    ws.append(["Item", "Part Name", "Path", "Quantity", "Thumbnail"])
    row = 2
    item_idx = 1
    for key, entry in bom_dict.items():
        name = entry["name"]
        path = entry["path"]
        qty = entry["qty"]
        ws.cell(row=row, column=1, value=item_idx)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=path)
        ws.cell(row=row, column=4, value=qty)
        # Insert thumbnail if exists
        thumb_file = os.path.join(thumbnail_folder, f"{assembly_name}_{item_idx}.png")
        if os.path.isfile(thumb_file):
            try:
                img = XLImage(thumb_file)
                img.width = 96
                img.height = 96
                img_anchor = f"E{row}"
                ws.add_image(img, img_anchor)
            except Exception:
                pass
        row += 1
        item_idx += 1


def main():
    # basic argument support: first arg = assembly folder, second = output xlsx
    if len(sys.argv) > 1:
        global ASSEMBLY_FOLDER
        ASSEMBLY_FOLDER = sys.argv[1]
    if len(sys.argv) > 2:
        global OUTPUT_XLSX
        OUTPUT_XLSX = sys.argv[2]

    sw = start_solidworks()
    wb = Workbook()
    # remove default sheet
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)

    assembly_files = [
        os.path.join(ASSEMBLY_FOLDER, f)
        for f in os.listdir(ASSEMBLY_FOLDER)
        if f.lower().endswith((".sldasm", ".sldasm_a", ".sldasm1"))
    ]
    assembly_files.sort()

    for assy_path in assembly_files:
        assy_name = os.path.splitext(os.path.basename(assy_path))[0]
        print("Opening assembly:", assy_path)
        try:
            model = sw.OpenDoc6(assy_path, swDocASSEMBLY, 0, "", 0, 0)
            # Wait briefly for SolidWorks to load components
            time.sleep(0.5)
            bom = gather_bom_from_assembly(model)
            # For each BOM entry, try to get a preview thumbnail
            idx = 1
            for key, entry in bom.items():
                comp = entry.get("component")
                thumb_path = os.path.join(THUMBNAIL_FOLDER, f"{assy_name}_{idx}.png")
                saved = False
                try:
                    model_doc = None
                    # Try to get model doc from component
                    try:
                        model_doc = comp.GetModelDoc2()
                    except Exception:
                        model_doc = None
                    # If we have a model doc, try saving a thumbnail image
                    if model_doc:
                        saved = save_thumbnail_from_model(model_doc, thumb_path)
                    else:
                        # Try opening the component file directly and saving an image
                        comp_path = entry.get("path")
                        if comp_path and os.path.isfile(comp_path):
                            try:
                                comp_doc = sw.OpenDoc6(comp_path, swDocPART, 0, "", 0, 0)
                                time.sleep(0.2)
                                saved = save_thumbnail_from_model(comp_doc, thumb_path)
                                try:
                                    sw.CloseDoc(comp_path)
                                except Exception:
                                    pass
                            except Exception:
                                saved = False
                except Exception:
                    saved = False
                idx += 1

            # Write BOM sheet
            write_bom_to_excel(bom, assy_name, wb, THUMBNAIL_FOLDER)
            # Close assembly
            try:
                sw.CloseDoc(assy_path)
            except Exception:
                pass
        except Exception as e:
            print("Failed to process", assy_path, e)
            traceback.print_exc()

    # Save workbook
    wb.save(OUTPUT_XLSX)
    print("Saved BOM workbook to", OUTPUT_XLSX)
    # Optionally quit SolidWorks
    try:
        # sw.ExitApp()  # uncomment to quit SolidWorks
        pass
    except Exception:
        pass


if __name__ == "__main__":
    main()
