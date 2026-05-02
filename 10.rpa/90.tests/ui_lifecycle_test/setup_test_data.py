# -*- coding: utf-8 -*-
"""
Test Data Setup
===============

테스트에 필요한 폴더 및 파일 구조를 생성합니다.
"""

import os
from pathlib import Path


def setup_test_data():
    """테스트 데이터 폴더 생성"""

    # BE 테스트 데이터
    be_folder = Path("D:/test_data/bom_exporter")
    be_folder.mkdir(parents=True, exist_ok=True)

    # 빈 DWG 파일 대신 설명 파일 생성
    readme = be_folder / "README.txt"
    with open(readme, "w", encoding="utf-8") as f:
        f.write("""BOM Exporter Test Data
======================

이 폴더에 테스트용 DWG 파일을 넣어주세요.

필요한 파일:
- BOM 속성이 포함된 DWG 파일들
- 다양한 블록과 속성을 포함하면 좋습니다

예시 파일 이름:
- test_drawing_01.dwg
- test_drawing_02.dwg
- sample_with_bom.dwg
""")

    print(f"[OK] BE test folder created: {be_folder}")

    # DC 테스트 데이터
    dc_folder = Path("D:/test_data/dwg_classifier/dwg_files")
    dc_folder.mkdir(parents=True, exist_ok=True)

    readme = dc_folder.parent / "README.txt"
    with open(readme, "w", encoding="utf-8") as f:
        f.write("""DWG Classifier Test Data
========================

이 폴더에 테스트용 DWG 파일과 분류 엑셀 파일을 넣어주세요.

폴더 구조:
  dwg_classifier/
  ├── dwg_files/          <- 분류할 DWG 파일들
  │   ├── drawing_01.dwg
  │   ├── drawing_02.dwg
  │   └── ...
  └── classification_list.xlsx  <- 분류 기준 엑셀 파일

엑셀 파일 형식:
- 첫 번째 열: 도면 번호 또는 파일명 패턴
- 두 번째 열: 분류 카테고리
""")

    # 샘플 엑셀 파일 생성 (openpyxl 필요)
    try:
        import openpyxl
        excel_path = dc_folder.parent / "classification_list.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "분류기준"

        # 헤더
        ws["A1"] = "도면번호"
        ws["B1"] = "분류"

        # 샘플 데이터
        sample_data = [
            ("A-001", "건축"),
            ("S-001", "구조"),
            ("M-001", "기계"),
            ("E-001", "전기"),
            ("P-001", "배관"),
        ]
        for i, (num, cat) in enumerate(sample_data, 2):
            ws[f"A{i}"] = num
            ws[f"B{i}"] = cat

        wb.save(str(excel_path))
        print(f"[OK] Sample Excel created: {excel_path}")

    except ImportError:
        print("  (openpyxl not installed, skipping Excel creation)")

    print(f"[OK] DC test folder created: {dc_folder}")

    # 결과 폴더
    results_folder = Path("D:/test_data/test_results")
    results_folder.mkdir(parents=True, exist_ok=True)
    print(f"[OK] Results folder created: {results_folder}")

    print("\n테스트 데이터 설정 완료!")
    print("\n다음 단계:")
    print("1. 테스트 폴더에 실제 DWG 파일을 복사하세요")
    print("2. DC 테스트를 위해 분류 엑셀 파일을 수정하세요")
    print("3. python run_all_tests.py 로 테스트를 실행하세요")


if __name__ == "__main__":
    setup_test_data()
