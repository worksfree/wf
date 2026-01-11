# -*- coding: utf-8 -*-
"""
K-Startup Web Automation
웹 자동화 핵심 로직과 Non-UI 모드 실행 담당
"""

import sys
import os
import argparse
import logging
from pathlib import Path
from datetime import datetime

# 경로 설정
current_dir = Path(__file__).resolve().parent
sys.path.insert(0, str(current_dir))

# 공통 유틸 경로
utils_path = current_dir.parent.parent.parent / "10.common"
if utils_path.exists():
    sys.path.append(str(utils_path))

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import Select, WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
    import openpyxl
    from openpyxl import load_workbook
except ImportError as e:
    print(f"필수 라이브러리 설치 필요: {e}")
    print("pip install selenium openpyxl 실행하세요")
    sys.exit(1)

import json
import time
import pickle
from typing import Dict, List, Any, Optional


class WebAutomationConfig:
    """설정 파일 관리"""
    
    def __init__(self, config_path: Optional[Path] = None):
        if config_path is None:
            # 기본 경로: ~/.wf_rpa/kstartup-web/settings.json
            config_path = Path.home() / ".wf_rpa" / "kstartup-web" / "settings.json"
        
        self.config_path = Path(config_path)
        self.config = self._load_config()
    
    def _load_config(self) -> Dict:
        """JSON 설정 파일 로드"""
        try:
            if self.config_path.exists():
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logging.warning(f"설정 파일 로드 실패: {e}")
        
        # 기본값
        return {
            "app_config": {
                "run_mode": "demo",
                "chrome_driver_path": "chromedriver",
                "headless_mode": False,
                "wait_timeout": 10,
                "screenshot_enabled": True
            },
            "ui_config": {
                "last_excel_path": ""
            },
            "logging_config": {
                "level": "INFO"
            }
        }
    
    def get(self, section: str, key: str, default=None):
        """설정값 조회"""
        if section in self.config and key in self.config[section]:
            return self.config[section][key]
        return default
    
    def set(self, section: str, key: str, value: Any):
        """설정값 저장"""
        if section not in self.config:
            self.config[section] = {}
        self.config[section][key] = value
        self._save_config()
    
    def _save_config(self):
        """설정값을 파일에 저장"""
        try:
            self.config_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logging.warning(f"설정 파일 저장 실패: {e}")


class WebActionParser:
    """Excel 파일에서 웹 액션 파싱"""
    
    # 지원하는 액션 타입
    SUPPORTED_ACTIONS = {
        'click': '요소 클릭',
        'input': '텍스트 입력',
        'select': '드롭다운 선택',
        'wait': '대기 (초)',
        'wait_element': '요소 대기',
        'scroll': '스크롤',
        'switch_window': '윈도우 전환',
        'get_text': '텍스트 추출',
        'check_element': '요소 확인',
        'navigate': '페이지 이동',
        'login': '로그인 실행'
    }
    
    def __init__(self, excel_path: str):
        """
        Args:
            excel_path: Excel 파일 경로
        """
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel 파일을 찾을 수 없습니다: {excel_path}")
        self.credentials = {}
    
    def load_credentials(self) -> Dict[str, str]:
        """
        Excel 파일의 'Config' 시트에서 인증 정보 로드
        
        Returns:
            credentials: {'login_id': ..., 'login_pw': ...}
        """
        try:
            wb = load_workbook(self.excel_path)
            
            # 'Config' 시트 확인
            if 'Config' in wb.sheetnames:
                ws = wb['Config']
                credentials = {}
                
                # 행 단위로 읽기 (키-값 쌍)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:  # 키와 값이 모두 있는 경우
                        key = str(row[0]).strip()
                        value = str(row[1]).strip()
                        credentials[key] = value
                
                self.credentials = credentials
                logging.info(f"인증 정보 로드 완료: {list(credentials.keys())}")
                return credentials
            else:
                logging.warning("'Config' 시트를 찾을 수 없습니다. 인증 정보를 건너뜁니다.")
                return {}
        
        except Exception as e:
            logging.error(f"인증 정보 로드 오류: {e}")
            return {}
    
    def parse(self, start_order: Optional[int] = None, end_order: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Excel 파일 파싱
        
        Args:
            start_order: 시작 순번 (None이면 전체)
            end_order: 종료 순번 (None이면 끝까지)
        
        컬럼 구성:
        - A: 순번 (번호)
        - B: 액션명 (action_name)
        - C: 액션 타입 (type) - click, input, select, wait 등
        - D: XPath/CSS Selector (xpath)
        - E: 입력값/파라미터 (value)
        - F: 대기시간 (wait_time, 옵션)
        - G: 설명 (description, 옵션)
        """
        actions = []
        try:
            wb = load_workbook(self.excel_path)
            
            # 'Actions' 시트 사용
            if 'Actions' in wb.sheetnames:
                ws = wb['Actions']
            else:
                ws = wb.active
            
            # 헤더 행 건너뛰기 (첫 행)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                # 빈 행 스킵
                if not any(cell.value for cell in row[:3]):
                    continue
                
                order = row[0].value
                
                # 순번 필터링
                if start_order is not None and (order is None or order < start_order):
                    continue
                if end_order is not None and (order is None or order > end_order):
                    continue
                
                action = {
                    'order': order,
                    'action_name': row[1].value,
                    'type': str(row[2].value).lower() if row[2].value else '',
                    'xpath': row[3].value or '',
                    'value': row[4].value or '',
                    'wait_time': row[5].value or 0,
                    'description': row[6].value or ''
                }
                
                # 액션 타입 검증
                if action['type'] not in self.SUPPORTED_ACTIONS:
                    logging.warning(f"지원하지 않는 액션 타입 ({row_idx}): {action['type']}")
                    continue
                
                actions.append(action)
            
            if start_order or end_order:
                logging.info(f"순번 {start_order or '시작'}~{end_order or '끝'}: {len(actions)}개의 액션을 파싱했습니다")
            else:
                logging.info(f"총 {len(actions)}개의 액션을 파싱했습니다")
            return actions
        
        except Exception as e:
            logging.error(f"Excel 파싱 오류: {e}")
            raise


class WebAutomationEngine:
    """Selenium 기반 웹 자동화 엔진"""
    
    def __init__(self, config: WebAutomationConfig):
        self.config = config
        self.driver = None
        self.logger = self._setup_logger()
        self.action_count = 0
        self.success_count = 0
        self.error_count = 0
        self.credentials = {}
        
        # 세션 쿠키 저장 경로
        self.cookie_dir = Path.home() / ".wf_rpa" / "kstartup-web" / "cookies"
        self.cookie_dir.mkdir(parents=True, exist_ok=True)
        self.cookie_file = self.cookie_dir / "session_cookies.pkl"

    
    def _setup_logger(self) -> logging.Logger:
        """로거 설정"""
        logger = logging.getLogger('WebAutomation')
        logger.setLevel(logging.INFO)
        
        # 파일 핸들러
        log_dir = Path.home() / ".wf_rpa" / "kstartup-web" / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)
        
        log_file = log_dir / f"automation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        fh = logging.FileHandler(log_file, encoding='utf-8')
        fh.setLevel(logging.INFO)
        
        # 포매터
        formatter = logging.Formatter(
            '[%(asctime)s] %(levelname)s: %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        
        return logger
    
    def initialize_driver(self) -> webdriver.Chrome:
        """Chrome WebDriver 초기화"""
        try:
            chrome_options = Options()
            
            if self.config.get('app_config', 'headless_mode', False):
                chrome_options.add_argument("--headless")
            
            # 창 크기 설정
            chrome_options.add_argument("--window-size=1920,1080")
            
            # 기타 옵션
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            
            driver_path = self.config.get('app_config', 'chrome_driver_path', 'chromedriver')
            # Selenium 4.x는 service와 options를 분리해야 함
            from selenium.webdriver.chrome.service import Service
            service = Service(driver_path)
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            self.logger.info("Chrome WebDriver 초기화 완료")
            return self.driver
        
        except Exception as e:
            self.logger.error(f"WebDriver 초기화 실패: {e}")
            raise
    
    def save_cookies(self, site_name: str = "kstartup"):
        """현재 세션의 쿠키를 파일에 저장"""
        try:
            cookies = self.driver.get_cookies()
            cookie_file = self.cookie_dir / f"{site_name}_cookies.pkl"
            
            with open(cookie_file, 'wb') as f:
                pickle.dump(cookies, f)
            
            self.logger.info(f"쿠키 저장 완료: {cookie_file}")
            return True
        except Exception as e:
            self.logger.error(f"쿠키 저장 실패: {e}")
            return False
    
    def load_cookies(self, site_name: str = "kstartup", url: str = None):
        """저장된 쿠키를 로드하여 세션 복원"""
        try:
            cookie_file = self.cookie_dir / f"{site_name}_cookies.pkl"
            
            if not cookie_file.exists():
                self.logger.info("저장된 쿠키가 없습니다. 새로운 세션을 시작합니다.")
                return False
            
            # 쿠키를 로드하기 전에 해당 도메인으로 먼저 접속
            if url:
                self.driver.get(url)
            
            with open(cookie_file, 'rb') as f:
                cookies = pickle.load(f)
            
            # 쿠키 추가
            for cookie in cookies:
                try:
                    self.driver.add_cookie(cookie)
                except Exception as e:
                    self.logger.warning(f"쿠키 추가 실패: {cookie.get('name')} - {e}")
            
            self.logger.info(f"쿠키 로드 완료: {len(cookies)}개")
            
            # 페이지 새로고침하여 쿠키 적용
            if url:
                self.driver.get(url)
            
            return True
        
        except Exception as e:
            self.logger.error(f"쿠키 로드 실패: {e}")
            return False
    
    def set_credentials(self, credentials: Dict[str, str]):
        """인증 정보 설정"""
        self.credentials = credentials
        self.logger.info(f"인증 정보 설정: {list(credentials.keys())}")
    
    def execute_actions(self, actions: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        액션 목록 실행
        
        Returns:
            실행 결과 딕셔너리
        """
        if not self.driver:
            self.initialize_driver()
        
        results = {
            'total': len(actions),
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        for action in actions:
            try:
                self._execute_single_action(action)
                results['success'] += 1
                self.logger.info(f"✓ 액션 실행 성공: {action['action_name']}")
            
            except Exception as e:
                results['failed'] += 1
                error_msg = f"액션 실행 실패 ({action['action_name']}): {str(e)}"
                results['errors'].append(error_msg)
                self.logger.error(error_msg)
        
        self.logger.info(f"\n=== 실행 결과 ===")
        self.logger.info(f"총: {results['total']}개, 성공: {results['success']}개, 실패: {results['failed']}개")
        
        return results
    
    def _execute_single_action(self, action: Dict[str, Any]):
        """단일 액션 실행"""
        action_type = action['type']
        
        if action_type == 'navigate':
            self._action_navigate(action['value'])
        
        elif action_type == 'wait':
            self._action_wait(float(action['value']))
        
        elif action_type == 'wait_element':
            self._action_wait_element(action['xpath'], int(action['wait_time']))
        
        elif action_type == 'click':
            self._action_click(action['xpath'])
        
        elif action_type == 'input':
            self._action_input(action['xpath'], str(action['value']))
        
        elif action_type == 'select':
            self._action_select(action['xpath'], str(action['value']))
        
        elif action_type == 'scroll':
            self._action_scroll(action['value'])
        
        elif action_type == 'get_text':
            text = self._action_get_text(action['xpath'])
            self.logger.info(f"추출된 텍스트: {text}")
        
        elif action_type == 'check_element':
            self._action_check_element(action['xpath'])
        
        elif action_type == 'switch_window':
            self._action_switch_window(int(action['value']))
        
        elif action_type == 'login':
            self._action_login(action)
    
    def _action_navigate(self, url: str):
        """페이지 이동"""
        self.driver.get(url)
        self.logger.info(f"페이지 이동: {url}")
    
    def _action_wait(self, seconds: float):
        """지정된 시간만큼 대기"""
        import time
        time.sleep(seconds)
        self.logger.info(f"{seconds}초 대기")
    
    def _action_wait_element(self, xpath: str, timeout: int):
        """요소가 나타날 때까지 대기"""
        wait = WebDriverWait(self.driver, timeout)
        element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        self.logger.info(f"요소 대기 완료: {xpath}")
        return element
    
    def _action_click(self, xpath: str):
        """요소 클릭"""
        wait = WebDriverWait(self.driver, self.config.get('app_config', 'wait_timeout', 10))
        element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
        element.click()
        self.logger.info(f"클릭: {xpath}")
    
    def _action_input(self, xpath: str, text: str):
        """텍스트 입력"""
        element = self.driver.find_element(By.XPATH, xpath)
        element.clear()
        element.send_keys(text)
        self.logger.info(f"입력: {xpath} = {text}")
    
    def _action_select(self, xpath: str, value: str):
        """드롭다운 선택"""
        element = self.driver.find_element(By.XPATH, xpath)
        select = Select(element)
        select.select_by_value(value)
        self.logger.info(f"선택: {xpath} = {value}")
    
    def _action_scroll(self, direction: str):
        """페이지 스크롤"""
        if direction.lower() == 'down':
            self.driver.execute_script("window.scrollBy(0, 500);")
        elif direction.lower() == 'up':
            self.driver.execute_script("window.scrollBy(0, -500);")
        self.logger.info(f"스크롤: {direction}")
    
    def _action_get_text(self, xpath: str) -> str:
        """요소의 텍스트 추출"""
        element = self.driver.find_element(By.XPATH, xpath)
        text = element.text
        return text
    
    def _action_check_element(self, xpath: str) -> bool:
        """요소 존재 여부 확인"""
        try:
            self.driver.find_element(By.XPATH, xpath)
            return True
        except NoSuchElementException:
            return False
    
    def _action_switch_window(self, window_index: int):
        """윈도우 전환"""
        windows = self.driver.window_handles
        self.driver.switch_to.window(windows[window_index])
        self.logger.info(f"윈도우 전환: {window_index}")
    
    def _action_login(self, action: Dict[str, Any]):
        """로그인 실행
        
        value 필드에 'id_xpath|pw_xpath|submit_xpath' 형식으로 지정
        또는 credentials에서 login_id, login_pw 사용
        """
        try:
            # XPath 파싱
            xpaths = action['xpath'].split('|')
            if len(xpaths) < 3:
                self.logger.error("로그인 액션은 'id_xpath|pw_xpath|submit_xpath' 형식이 필요합니다")
                return
            
            id_xpath, pw_xpath, submit_xpath = xpaths[0].strip(), xpaths[1].strip(), xpaths[2].strip()
            
            # 인증 정보 가져오기
            login_id = self.credentials.get('login_id', '')
            login_pw = self.credentials.get('login_pw', '')
            
            if not login_id or not login_pw:
                self.logger.error("인증 정보(login_id, login_pw)가 설정되지 않았습니다")
                return
            
            # ID 입력
            id_element = self.driver.find_element(By.XPATH, id_xpath)
            id_element.clear()
            id_element.send_keys(login_id)
            self.logger.info(f"ID 입력: {login_id}")
            
            time.sleep(0.5)
            
            # 비밀번호 입력
            pw_element = self.driver.find_element(By.XPATH, pw_xpath)
            pw_element.clear()
            pw_element.send_keys(login_pw)
            self.logger.info("비밀번호 입력 완료")
            
            time.sleep(0.5)
            
            # 로그인 버튼 클릭
            submit_element = self.driver.find_element(By.XPATH, submit_xpath)
            submit_element.click()
            self.logger.info("로그인 버튼 클릭")
            
            # 로그인 후 대기
            time.sleep(2)
            
            # 로그인 성공 후 쿠키 저장
            self.save_cookies("kstartup")
            
        except Exception as e:
            self.logger.error(f"로그인 실패: {e}")
            raise
    
    def close_driver(self):
        """WebDriver 종료"""
        if self.driver:
            self.driver.quit()
            self.logger.info("WebDriver 종료")


def main():
    """Non-UI 모드: 명령행 실행"""
    parser = argparse.ArgumentParser(
        description='K-Startup 웹 자동화 엔진',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예제:
  python automation.py path/to/actions.xlsx
  python automation.py path/to/actions.xlsx --config ~/.wf_rpa/kstartup-web/settings.json
  python automation.py path/to/actions.xlsx --start 5 --end 10
  python automation.py path/to/actions.xlsx --use-session
        """
    )
    
    parser.add_argument('excel_file', help='웹 액션이 정의된 Excel 파일 경로')
    parser.add_argument(
        '--config',
        help='설정 파일 경로 (기본: ~/.wf_rpa/kstartup-web/settings.json)'
    )
    parser.add_argument(
        '--headless',
        action='store_true',
        help='Headless 모드 실행 (브라우저 창 숨김)'
    )
    parser.add_argument(
        '--start',
        type=int,
        help='시작 순번 (구간별 실행)'
    )
    parser.add_argument(
        '--end',
        type=int,
        help='종료 순번 (구간별 실행)'
    )
    parser.add_argument(
        '--use-session',
        action='store_true',
        help='저장된 세션 쿠키 사용 (로그인 생략)'
    )
    
    args = parser.parse_args()
    
    # 로깅 설정
    logging.basicConfig(
        level=logging.INFO,
        format='[%(asctime)s] %(levelname)s: %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    try:
        # 설정 로드
        config = WebAutomationConfig(args.config)
        
        if args.headless:
            config.set('app_config', 'headless_mode', True)
        
        # Excel 파싱
        logging.info(f"Excel 파일 로드: {args.excel_file}")
        parser_obj = WebActionParser(args.excel_file)
        
        # 인증 정보 로드
        credentials = parser_obj.load_credentials()
        
        # 액션 파싱 (구간 지정 가능)
        actions = parser_obj.parse(start_order=args.start, end_order=args.end)
        
        # 엔진 초기화 및 실행
        engine = WebAutomationEngine(config)
        engine.set_credentials(credentials)
        engine.initialize_driver()
        
        # 세션 복원 (옵션)
        if args.use_session:
            # K-Startup 메인 페이지로 이동 후 쿠키 로드
            engine.driver.get("https://www.k-startup.go.kr/")
            engine.load_cookies("kstartup", "https://www.k-startup.go.kr/")
            logging.info("저장된 세션을 사용합니다.")
        
        results = engine.execute_actions(actions)
        
        engine.close_driver()
        
        # 종료 코드 반환
        return 0 if results['failed'] == 0 else 1
    
    except Exception as e:
        logging.error(f"오류 발생: {e}")
        return 1


if __name__ == '__main__':
    sys.exit(main())
