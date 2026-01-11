# %% [셀 0] 임포트 및 초기 설정
# -*- coding: utf-8 -*-
"""
K-Startup 웹 자동화 - 간단한 인터랙티브 버전

사용법:
  1. 셀 0 실행 (Shift+Enter)
  2. 셀 1 실행 → Excel 로드
  3. 셀 2 실행 → Chrome 시작 + 로그인
  4. 셀 3 실행 → 첫 액션
  ...반복

주의: 셀 2 (로그인)는 한 번만!
"""

import sys
import time
import logging
import os
from pathlib import Path
from datetime import datetime

# UTF-8 인코딩
try:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
except:
    pass

# 라이브러리 임포트
try:
    from openpyxl import load_workbook
except:
    print("⚠️ openpyxl 필요: pip install openpyxl")

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.keys import Keys
except:
    print("⚠️ selenium 필요: pip install selenium")

try:
    import pyautogui
except:
    print("⚠️ pyautogui 필요: pip install pyautogui")

try:
    import pyperclip
except:
    print("⚠️ pyperclip 필요: pip install pyperclip")

# 로깅 설정
def setup_logger():
    logger = logging.getLogger("kstartup")
    logger.setLevel(logging.DEBUG)
    
    # 콘솔 핸들러
    handler = logging.StreamHandler(sys.stdout)
    handler.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(message)s', datefmt='%H:%M:%S')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    
    # 파일 핸들러
    log_dir = Path(__file__).parent / "logs"
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / f"kstartup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    
    return logger

logger = setup_logger()
logger.info("✓ 임포트 완료")


# %% [셀 1] Excel 파일 읽기
def load_excel():
    """Excel에서 Config와 Actions 데이터 읽기"""
    global config_data, actions_data
    
    logger.info("="*70)
    logger.info("Excel 파일 로드")
    logger.info("="*70)
    
    excel_path = Path(__file__).parent / "회계감사비.xlsx"
    
    if not excel_path.exists():
        logger.error(f"❌ Excel 파일 없음: {excel_path}")
        logger.error("대신 사용할 파일: 창업활동비.xlsx")
        excel_path = Path(__file__).parent / "창업활동비.xlsx"
        
        if not excel_path.exists():
            logger.error(f"❌ 두 파일 모두 없음!")
            return False
    
    try:
        workbook = load_workbook(excel_path)
        
        # Config 시트
        config_sheet = workbook['Config']
        config_data = {}
        for row in config_sheet.iter_rows(min_row=2, values_only=False):
            key_cell = row[0]
            val_cell = row[1]
            if key_cell.value:
                config_data[key_cell.value] = val_cell.value if val_cell.value else ""
        
        # Actions 시트
        actions_sheet = workbook['Actions']
        actions_data = []
        headers = []
        
        for row_idx, row in enumerate(actions_sheet.iter_rows(min_row=1, values_only=True), 1):
            if row_idx == 1:
                headers = [h for h in row if h]
            else:
                action_dict = {}
                for col_idx, header in enumerate(headers):
                    if col_idx < len(row):
                        action_dict[header] = row[col_idx]
                actions_data.append(action_dict)
        
        logger.info(f"✓ Excel 로드 완료")
        logger.info(f"  - Config: {len(config_data)} 항목")
        logger.info(f"  - Actions: {len(actions_data)} 개")
        logger.info(f"  - 로그인 ID: {config_data.get('login_id', '?')}")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Excel 읽기 실패: {e}")
        return False

# 전역 변수
config_data = None
actions_data = None
driver = None

load_excel()


# %% [셀 2] 로그인 (Chrome 시작)
def start_browser_and_login():
    """Chrome 시작 및 로그인"""
    global driver, config_data
    
    logger.info("="*70)
    logger.info("Chrome 시작 및 로그인")
    logger.info("="*70)
    
    if driver:
        logger.info("⚠️ 이미 Chrome이 실행 중입니다")
        return
    
    if not config_data:
        logger.error("❌ Excel 설정을 먼저 로드하세요 (셀 1 실행)")
        return
    
    try:
        # Chrome 옵션
        options = webdriver.ChromeOptions()
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        
        # Chrome 시작
        logger.info("Chrome 시작 중...")
        driver = webdriver.Chrome(options=options)
        logger.info("✓ Chrome 시작됨")
        
        # K-Startup 사이트로 이동
        url = config_data.get('base_url', 'https://www.k-startup.go.kr')
        logger.info(f"사이트 접속: {url}")
        driver.get(url)
        
        # 로그인 폼 대기
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "userEmail"))
        )
        logger.info("✓ 로그인 폼 로드됨")
        
        # ID 입력
        login_id = config_data.get('login_id', '')
        login_pw = config_data.get('login_pw', '')
        
        if not login_id or not login_pw:
            logger.error("❌ Excel의 Config에 login_id, login_pw가 없습니다")
            driver.quit()
            driver = None
            return
        
        logger.info(f"사용자명 입력: {login_id}")
        id_input = driver.find_element(By.NAME, "userEmail")
        id_input.clear()
        id_input.send_keys(login_id)
        time.sleep(0.3)
        
        # 비밀번호 입력
        logger.info("비밀번호 입력")
        pw_input = driver.find_element(By.NAME, "userPwd")
        pw_input.clear()
        pw_input.send_keys(login_pw)
        time.sleep(0.3)
        
        # 로그인 버튼 클릭
        logger.info("로그인 버튼 클릭...")
        submit = driver.find_element(By.XPATH, '//button[@type="submit"]')
        submit.click()
        
        # 로그인 완료 대기
        time.sleep(3)
        logger.info(f"✓ 로그인 완료: {driver.title}")
        
    except Exception as e:
        logger.error(f"❌ 로그인 실패: {e}")
        if driver:
            driver.quit()
            driver = None

# 실행
start_browser_and_login()


# %% [셀 3] 액션 1 (로그인 후 첫 번째)
logger.info("\n" + "="*70)
logger.info("액션 1 실행")
logger.info("="*70)

if driver and actions_data and len(actions_data) > 0:
    action = actions_data[0]
    logger.info(f"액션명: {action.get('액션명', '?')}")
    logger.info(f"URL: {action.get('URL', '?')}")
    logger.info(f"엘리먼트: {action.get('엘리먼트', '?')}")
    logger.info("✓ 현재 정보 확인됨 - 실행할 준비 완료")
else:
    logger.error("❌ Chrome이나 액션 데이터 없음")


# %% [셀 4] 현재 상태 확인
logger.info("\n" + "="*70)
logger.info("현재 상태")
logger.info("="*70)

logger.info(f"Chrome 실행 여부: {'✓ Yes' if driver else '❌ No'}")
logger.info(f"Excel 로드 여부: {'✓ Yes' if config_data else '❌ No'}")
logger.info(f"액션 개수: {len(actions_data) if actions_data else 0}")

if driver:
    logger.info(f"현재 페이지: {driver.title}")
    logger.info(f"현재 URL: {driver.current_url}")


# %% [셀 5] 브라우저 종료
if driver:
    logger.info("\n브라우저 종료...")
    driver.quit()
    driver = None
    logger.info("✓ 브라우저 종료됨")
