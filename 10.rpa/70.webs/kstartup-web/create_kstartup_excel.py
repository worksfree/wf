# -*- coding: utf-8 -*-
"""
K-Startup 로그인 샘플 Excel 파일 생성
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

def create_kstartup_excel():
    """K-Startup 로그인 샘플 Excel 파일 생성"""
    
    wb = Workbook()
    
    # ========== Config 시트 생성 ==========
    ws_config = wb.active
    ws_config.title = "Config"
    
    # Config 헤더
    config_headers = ["Key", "Value", "설명"]
    ws_config.append(config_headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws_config[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Config 데이터
    config_data = [
        ["login_id", "tanklis", "K-Startup 로그인 ID"],
        ["login_pw", "lis@9154", "K-Startup 로그인 비밀번호"],
        ["site_url", "https://www.k-startup.go.kr/", "K-Startup 사이트 URL"],
    ]
    
    for row in config_data:
        ws_config.append(row)
    
    # Config 컬럼 너비
    ws_config.column_dimensions['A'].width = 20
    ws_config.column_dimensions['B'].width = 30
    ws_config.column_dimensions['C'].width = 40
    
    # ========== Actions 시트 생성 ==========
    ws_actions = wb.create_sheet("Actions")
    
    # Actions 헤더
    actions_headers = ["순번", "액션명", "액션 타입", "XPath/Selector", "입력값/파라미터", "대기시간(초)", "설명"]
    ws_actions.append(actions_headers)
    
    # 헤더 스타일
    action_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for cell in ws_actions[1]:
        cell.fill = action_header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Actions 샘플 데이터
    # 주의: K-Startup 로그인 XPath는 실제 사이트 확인 후 수정 필요
    actions_data = [
        [1, "K-Startup 메인페이지 접속", "navigate", "", "https://www.k-startup.go.kr/", "", "메인페이지 이동"],
        [2, "페이지 로드 대기", "wait", "", "2", "", "페이지 로딩 대기"],
        [3, "로그인 버튼 찾기 및 클릭", "click", "//a[contains(text(), '로그인')]", "", "3", "로그인 페이지로 이동"],
        [4, "로그인 페이지 대기", "wait", "", "2", "", "로그인 페이지 로딩"],
        [5, "로그인 실행", "login", "//input[@id='user_id']|//input[@id='user_pw']|//button[@type='submit']", "", "", "ID/PW 입력 및 로그인 (Config에서 읽음)"],
        [6, "로그인 후 대기", "wait", "", "3", "", "로그인 처리 대기"],
        [7, "메인 페이지 확인", "check_element", "//div[@class='user-info']", "", "", "로그인 성공 확인"],
    ]
    
    for row in actions_data:
        ws_actions.append(row)
    
    # Actions 컬럼 너비
    column_widths = [8, 20, 15, 50, 25, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws_actions.column_dimensions[chr(64 + i)].width = width
    
    # 테두리 스타일 적용
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for ws in [ws_config, ws_actions]:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    # 파일 저장
    output_path = Path(__file__).parent / "kstartup_actions.xlsx"
    
    # 이미 파일이 있으면 건너뛰기
    if output_path.exists():
        print(f"⚠️  Excel 파일이 이미 존재합니다: {output_path}")
        print(f"   기존 파일을 템플릿으로 사용하세요.")
        print(f"   새로 생성하려면 먼저 기존 파일을 삭제하세요.")
        return output_path
    
    wb.save(output_path)
    print(f"✓ K-Startup 템플릿 Excel 파일 생성 완료: {output_path}")
    print(f"  - Config 시트: 로그인 정보 포함")
    print(f"  - Actions 시트: {len(actions_data)}개 액션 정의")
    print()
    print("주의사항:")
    print("1. Actions 시트의 XPath는 실제 K-Startup 사이트에 맞게 수정해야 합니다.")
    print("2. 로그인 테스트 전에 실제 사이트의 요소를 개발자 도구로 확인하세요.")
    print("3. Config 시트의 login_id, login_pw를 실제 계정 정보로 수정하세요.")
    
    return output_path

if __name__ == "__main__":
    create_kstartup_excel()
