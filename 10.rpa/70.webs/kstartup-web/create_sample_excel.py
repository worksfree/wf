# -*- coding: utf-8 -*-
"""
K-Startup 웹 자동화 샘플 Excel 파일 생성
Config 시트와 Actions 시트를 포함한 템플릿 생성
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

def create_sample_excel():
    """샘플 엑셀 파일 생성"""
    
    # 워크북 생성
    wb = Workbook()
    
    # 기본 시트 제거
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========== Config 시트 ==========
    ws_config = wb.create_sheet('Config', 0)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # 헤더 설정
    ws_config['A1'] = '키'
    ws_config['B1'] = '값'
    ws_config['A1'].fill = header_fill
    ws_config['A1'].font = header_font
    ws_config['B1'].fill = header_fill
    ws_config['B1'].font = header_font
    ws_config['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_config['B1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Config 데이터
    config_data = [
        ['url', 'https://www.k-startup.go.kr'],
        ['login_id', '사용자ID'],
        ['login_pw', '사용자비밀번호'],
        ['', ''],
        ['# 주석: Config 시트에는 URL, ID, PW 등 설정값을 입력합니다', '']
    ]
    
    for row_idx, (key, value) in enumerate(config_data, start=2):
        ws_config[f'A{row_idx}'] = key
        ws_config[f'B{row_idx}'] = value
        
        # 주석 행 스타일
        if key.startswith('#'):
            ws_config[f'A{row_idx}'].font = Font(italic=True, color="808080")
    
    # 열 너비 조정
    ws_config.column_dimensions['A'].width = 30
    ws_config.column_dimensions['B'].width = 40
    
    # ========== Actions 시트 ==========
    ws_actions = wb.create_sheet('Actions', 1)
    
    # 헤더
    headers = ['순번', '액션명', '타입', 'XPath', '값', '대기시간', '설명']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_actions.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 샘플 액션 데이터
    actions_data = [
        [1, '페이지 이동', 'navigate', '', 'https://www.k-startup.go.kr', 2, 'K-Startup 메인 페이지로 이동'],
        [2, '로그인 버튼 클릭', 'click', '//a[contains(@class, "login")]', '', 1, '로그인 페이지로 이동'],
        [3, '아이디 입력', 'input', '//input[@id="userId"]', '{login_id}', 0.5, 'Config 시트의 login_id 사용'],
        [4, '비밀번호 입력', 'input', '//input[@id="password"]', '{login_pw}', 0.5, 'Config 시트의 login_pw 사용'],
        [5, '로그인 실행', 'click', '//button[@type="submit"]', '', 2, '로그인 버튼 클릭'],
        [6, '페이지 대기', 'wait', '', '3', 3, '페이지 로딩 대기 (초)'],
        ['', '', '', '', '', '', ''],
        ['', '# 지원 액션 타입:', '', '', '', '', ''],
        ['', '# - navigate: 페이지 이동 (값에 URL)', '', '', '', '', ''],
        ['', '# - click: 요소 클릭 (XPath)', '', '', '', '', ''],
        ['', '# - input: 텍스트 입력 (XPath + 값)', '', '', '', '', ''],
        ['', '# - select: 드롭다운 선택 (XPath + 값)', '', '', '', '', ''],
        ['', '# - wait: 대기 (값에 초)', '', '', '', '', ''],
        ['', '# - wait_element: 요소 대기 (XPath)', '', '', '', '', ''],
        ['', '# - get_text: 텍스트 추출 (XPath)', '', '', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['', '# Config 변수 사용: {키} 형태로 사용', '', '', '', '', ''],
        ['', '# 예: {login_id}, {login_pw}, {url}', '', '', '', '', ''],
    ]
    
    # 데이터 입력
    for row_idx, row_data in enumerate(actions_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_actions.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # 주석 행 스타일
            if isinstance(value, str) and value.startswith('#'):
                cell.font = Font(italic=True, color="808080")
            
            # 일반 데이터 정렬
            if col_idx == 1:  # 순번
                cell.alignment = Alignment(horizontal='center')
            elif col_idx in [5, 6]:  # 값, 대기시간
                cell.alignment = Alignment(horizontal='right')
    
    # 열 너비 조정
    ws_actions.column_dimensions['A'].width = 8   # 순번
    ws_actions.column_dimensions['B'].width = 20  # 액션명
    ws_actions.column_dimensions['C'].width = 15  # 타입
    ws_actions.column_dimensions['D'].width = 35  # XPath
    ws_actions.column_dimensions['E'].width = 30  # 값
    ws_actions.column_dimensions['F'].width = 10  # 대기시간
    ws_actions.column_dimensions['G'].width = 40  # 설명
    
    # 테두리 적용
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for ws in [ws_config, ws_actions]:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    # 파일 저장
    output_path = Path(__file__).parent / 'kstartup_actions.xlsx'
    wb.save(output_path)
    print(f"✓ 샘플 엑셀 파일 생성 완료: {output_path}")
    print(f"  - Config 시트: 설정값 (URL, ID, PW)")
    print(f"  - Actions 시트: 자동화 액션 정의")
    
    return output_path


if __name__ == '__main__':
    create_sample_excel()
