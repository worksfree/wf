# %% 임포트 및 초기 설정 (맨 처음 1회 실행 필수!)
# -*- coding: utf-8 -*-
"""
K-Startup 웹 자동화 - 단순 테스트 버전
단계 1: Excel 파일 읽기
단계 2: 사이트 접속
"""

# 임포트 및 설정
import sys
import time
import random
import logging
import os
import socket
from pathlib import Path
from datetime import datetime

# UTF-8 콘솔 인코딩 설정 (Windows 호환, Jupyter 호환)
try:
    if hasattr(sys.stdout, 'reconfigure') and sys.stdout.encoding != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, 'reconfigure') and sys.stderr.encoding != 'utf-8':
        sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass  # Jupyter/IPython 환경에서는 건너뜀

# openpyxl로 Excel 읽기
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl이 설치되지 않았습니다. pip install openpyxl")

# Selenium으로 브라우저 제어
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import Select
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False
    print("⚠️ selenium이 설치되지 않았습니다. pip install selenium")

# 화면 기반 클릭을 위한 pyautogui (선택)
try:
    import pyautogui
    PYAUTO_AVAILABLE = True
    try:
        import cv2  # confidence 매칭 향상용
        PYAUTO_OPENCV = True
    except ImportError:
        PYAUTO_OPENCV = False
    pyautogui.FAILSAFE = False
except ImportError:
    PYAUTO_AVAILABLE = False
    PYAUTO_OPENCV = False
    print("⚠️ pyautogui가 설치되지 않았습니다. pip install pyautogui opencv-python")

# 클립보드 관리 (파일 업로드용)
try:
    import pyperclip
    PYPERCLIP_AVAILABLE = True
except ImportError:
    PYPERCLIP_AVAILABLE = False
    print("⚠️ pyperclip이 설치되지 않았습니다. pip install pyperclip")

# Timeout 예외 처리용
from selenium.common.exceptions import (
    TimeoutException, 
    ElementClickInterceptedException,
    StaleElementReferenceException,
    NoSuchElementException
)


# 로깅 설정
def setup_logger(name="kstartup_simple"):
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)
    
    # 콘솔 핸들러
    handler = logging.StreamHandler()
    handler.setLevel(logging.INFO)
    formatter = logging.Formatter('[%(asctime)s] %(levelname)s: %(message)s', datefmt='%H:%M:%S')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    
    return logger


logger = setup_logger()


# ============================================================================
# 사용자 입력 대기 함수
# ============================================================================
def wait_for_user_input(message=""):
    """터미널에서 사용자 입력을 대기합니다.
    
    Args:
        message (str): 출력할 메시지
    """
    if message:
        print(f"\n{message}")
    input("\n→ 아무 키나 누르세요...")
    logger.info("✓ 사용자 입력 받음. 계속 진행...\n")


# ============================================================================
# 액션 에러 처리 공통 함수
# ============================================================================
def handle_action_error(exception, action_type, element_info="", retry_func=None, retry_args=None):
    """액션 실행 중 발생한 에러를 처리하고 사용자 입력을 받아 재시도합니다.
    
    Args:
        exception: 발생한 예외 객체
        action_type: 액션 타입 (예: 'click', 'input')
        element_info: 요소 정보 (XPath 등)
        retry_func: 재시도할 함수 (선택)
        retry_args: 재시도 함수의 인자 (튜플 또는 딕셔너리)
    
    Returns:
        재시도 성공 시 True, 실패 시 예외 발생
    """
    error_type = type(exception).__name__
    
    # 에러 유형별 메시지 정의
    error_messages = {
        'ElementClickInterceptedException': {
            'title': '요소 클릭 차단됨 (Element Click Intercepted)',
            'reason': '다른 요소가 클릭 대상을 가리고 있습니다.',
            'cause': '팝업, 모달, 오버레이, 또는 로딩 중인 요소',
            'solution': '팝업을 닫거나 요소가 로드될 때까지 대기',
            'tips': [
                '페이지의 팝업/모달을 수동으로 닫아주세요',
                '혹은 페이지가 완전히 로드될 때까지 기다리세요',
                '준비 완료되면 터미널에서 아무 키나 누르세요'
            ]
        },
        'TimeoutException': {
            'title': '요소 찾기 시간 초과 (Timeout)',
            'reason': '지정된 시간 내에 요소를 찾지 못했습니다.',
            'cause': '요소가 아직 로드되지 않았거나, XPath/CSS가 잘못됨, iframe 컨텍스트 오류',
            'solution': '페이지 로드 완료 확인 또는 XPath/CSS 수정',
            'tips': [
                '페이지가 완전히 로드될 때까지 기다리세요',
                '요소가 올바른 iframe 안에 있는지 확인하세요',
                'XPath/CSS 선택자가 정확한지 확인하세요',
                '준비 완료되면 터미널에서 아무 키나 누르세요'
            ]
        },
        'StaleElementReferenceException': {
            'title': '요소 참조 만료 (Stale Element)',
            'reason': 'DOM이 변경되어 이전에 찾은 요소가 더 이상 유효하지 않습니다.',
            'cause': 'JavaScript가 페이지를 동적으로 재구성함',
            'solution': '페이지 변경 완료 후 요소를 다시 찾음',
            'tips': [
                '페이지의 동적 변경이 완료될 때까지 기다리세요',
                '준비 완료되면 터미널에서 아무 키나 누르세요'
            ]
        },
        'NoSuchElementException': {
            'title': '요소를 찾을 수 없음 (No Such Element)',
            'reason': '지정된 선택자에 해당하는 요소가 존재하지 않습니다.',
            'cause': 'XPath/CSS 선택자가 잘못되었거나, iframe 컨텍스트 오류',
            'solution': 'XPath/CSS 확인 또는 iframe 전환 확인',
            'tips': [
                'XPath/CSS 선택자가 정확한지 확인하세요',
                '올바른 iframe 안에 있는지 확인하세요',
                '준비 완료되면 터미널에서 아무 키나 누르세요'
            ]
        }
    }
    
    # 기본 에러 메시지
    default_error = {
        'title': f'액션 실행 오류 ({error_type})',
        'reason': f'{action_type} 액션 실행 중 에러가 발생했습니다.',
        'cause': str(exception),
        'solution': '문제를 수동으로 해결하고 계속 진행',
        'tips': [
            '브라우저 상태를 확인하세요',
            '문제를 수동으로 해결하세요',
            '준비 완료되면 터미널에서 아무 키나 누르세요'
        ]
    }
    
    error_info = error_messages.get(error_type, default_error)
    
    # 에러 정보 출력
    logger.error(f"\n❌ {error_info['title']}")
    logger.error(f"   → {error_info['reason']}")
    logger.error(f"   → 원인: {error_info['cause']}")
    logger.error(f"   → 해결: {error_info['solution']}")
    if element_info:
        logger.error(f"   → 요소: {element_info}")

    # 예외 시 창 핸들 상태도 점검 (창 변경으로 인한 인터셉트 대응)
    try:
        global driver
        handles = driver.window_handles if 'driver' in globals() and driver else []
        if handles:
            curr_handle = driver.current_window_handle
            logger.info(f"창 핸들 상태: 총 {len(handles)}개, 현재: {curr_handle}")
            switch_to_latest_window(driver)
            logger.info(f"창 핸들(전환 후): {driver.current_window_handle}")
    except Exception as e:
        logger.warning(f"⚠️ 창 핸들 점검 중 예외: {e}")
    
    # 팁 메시지 구성
    tip_message = "\n💡 팁:"
    for i, tip in enumerate(error_info['tips'], 1):
        tip_message += f"\n   {i}. {tip}"
    
    # 사용자 입력 대기
    wait_for_user_input(tip_message)
    
    # 재시도 함수가 제공된 경우 실행
    if retry_func:
        logger.info(f"→ {action_type} 액션 재시도 중...")
        try:
            if isinstance(retry_args, dict):
                return retry_func(**retry_args)
            elif isinstance(retry_args, tuple):
                return retry_func(*retry_args)
            else:
                return retry_func()
        except Exception as retry_error:
            logger.error(f"❌ 재시도 실패: {retry_error}")
            raise
    
    return True


# ============================================================================
# 창/탭 포커스 헬퍼
# ============================================================================
def switch_to_latest_window(driver):
    """새 창/탭이 열렸으면 가장 최근 창으로 포커스를 전환합니다. (페이지 로드 대기 포함)"""
    try:
        handles = driver.window_handles
        if not handles:
            return
        latest = handles[-1]
        if driver.current_window_handle != latest:
            driver.switch_to.window(latest)
            logger.info(f"✓ 최신 창으로 전환: {latest[:20]}...")
            
            # 페이지 로드 완료 대기 (중요!)
            time.sleep(1)  # 1초 대기
            try:
                WebDriverWait(driver, 10).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
                logger.info(f"✓ 페이지 로드 완료: {driver.current_url}")
            except TimeoutException:
                logger.warning(f"⚠️ 페이지 로드 대기 시간 초과: {driver.current_url}")
    except Exception as e:
        logger.warning(f"⚠️ 창 전환 중 예외: {e}")


def is_logged_in(driver):
    """페이지에 로그아웃/마이페이지 등 로그인 전용 요소가 있는지 검사합니다."""
    selectors = [
        (By.XPATH, "//a[contains(@href,'logout') or contains(text(),'로그아웃')]") ,
        (By.XPATH, "//*[@id='header']//a[contains(@href,'logout')]") ,
        (By.XPATH, "//button[contains(text(),'로그아웃')]") ,
        (By.XPATH, "//a[contains(text(),'마이페이지') or contains(text(),'내정보')]") ,
    ]
    for by, locator in selectors:
        try:
            el = driver.find_element(by, locator)
            if el.is_displayed():
                logger.info("✓ 로그인 상태 감지 (로그아웃/마이페이지 요소 확인)")
                return True
        except Exception:
            continue

    # 로그인 버튼이 보이면 비로그인으로 판단
    try:
        login_btn = driver.find_element(By.XPATH, "//a[contains(text(),'로그인') or contains(@href,'login')]")
        if login_btn.is_displayed():
            logger.info("ℹ️ 로그인 버튼 감지 → 비로그인 상태")
            return False
    except Exception:
        pass

    logger.info("ℹ️ 로그인 상태를 확정할 수 없음 (기본값: 미로그인)")
    return False


# ============================================================================
# 요소 찾기 진단 함수
# ============================================================================
def diagnose_element_issue(driver, xpath):
    """요소를 찾지 못할 때 원인을 진단합니다."""
    logger.info(f"\n🔍 요소 찾기 진단: {xpath}")
    logger.info("-" * 70)
    
    # 1. 현재 창 정보
    try:
        logger.info(f"현재 URL: {driver.current_url}")
        logger.info(f"현재 페이지 제목: {driver.title}")
        logger.info(f"창 핸들 개수: {len(driver.window_handles)}")
    except Exception as e:
        logger.warning(f"⚠️ 창 정보 조회 실패: {e}")
    
    # 2. iframe 정보
    try:
        driver.switch_to.default_content()
        iframes = driver.find_elements(By.TAG_NAME, 'iframe')
        logger.info(f"현재 페이지 iframe 개수: {len(iframes)}")
        for i, iframe in enumerate(iframes, 1):
            iframe_id = iframe.get_attribute('id') or '(id 없음)'
            iframe_name = iframe.get_attribute('name') or '(name 없음)'
            logger.info(f"  [{i}] id='{iframe_id}' name='{iframe_name}'")
    except Exception as e:
        logger.warning(f"⚠️ iframe 정보 조회 실패: {e}")
    
    # 3. 요소가 DOM에 있는지 확인 (완전히 로드되지 않았거나 숨겨졌을 수 있음)
    try:
        elements = driver.find_elements(By.XPATH, xpath)
        if elements:
            logger.info(f"✓ 요소 발견: {len(elements)}개")
            
            for i, elem in enumerate(elements, 1):
                try:
                    is_displayed = elem.is_displayed()
                    is_enabled = elem.is_enabled()
                    tag = elem.tag_name
                    text = elem.text[:50] if elem.text else '(텍스트 없음)'
                    logger.info(f"  [{i}] {tag} | 표시: {is_displayed} | 활성: {is_enabled} | 텍스트: {text}")
                except Exception as e:
                    logger.warning(f"  [{i}] 요소 정보 조회 실패: {e}")
        else:
            logger.warning(f"❌ XPath에 일치하는 요소가 없음: {xpath}")
            logger.info("💡 시도:")
            logger.info("   1. XPath가 정확한지 개발자 도구(F12)에서 확인")
            logger.info("   2. 요소가 iframe 안에 있는지 확인")
            logger.info("   3. JavaScript 콘솔에서 $x('{}') 실행해보기".format(xpath))
    except Exception as e:
        logger.error(f"❌ 요소 검색 오류: {e}")
    
    # 4. 페이지 소스 길이 (로드 여부 판단)
    try:
        source_len = len(driver.page_source)
        logger.info(f"페이지 소스 길이: {source_len} bytes")
        if source_len < 1000:
            logger.warning("⚠️ 페이지가 제대로 로드되지 않은 것 같습니다")
    except Exception as e:
        logger.warning(f"⚠️ 페이지 소스 조회 실패: {e}")
    
    logger.info("-" * 70 + "\n")


def list_all_elements(driver, tag_name="a"):
    """현재 페이지의 특정 태그 모든 요소를 나열합니다."""
    try:
        driver.switch_to.default_content()
        elements = driver.find_elements(By.TAG_NAME, tag_name)
        logger.info(f"\n📋 모든 '{tag_name}' 요소 ({len(elements)}개):")
        logger.info("-" * 70)
        for i, elem in enumerate(elements[:20], 1):  # 처음 20개만
            try:
                text = elem.text[:40] if elem.text else '(텍스트 없음)'
                href = elem.get_attribute('href') or '(href 없음)'
                is_displayed = elem.is_displayed()
                logger.info(f"[{i}] {text} | href={href[:50]} | 표시={is_displayed}")
            except Exception as e:
                logger.warning(f"[{i}] 요소 정보 조회 실패: {e}")
        if len(elements) > 20:
            logger.info(f"... 외 {len(elements) - 20}개")
        logger.info("-" * 70 + "\n")
    except Exception as e:
        logger.error(f"❌ 요소 나열 실패: {e}")


def get_page_info(driver):
    """현재 페이지의 상세 정보를 출력합니다."""
    logger.info("\n📊 페이지 상세 정보:")
    logger.info("=" * 70)
    try:
        logger.info(f"URL: {driver.current_url}")
        logger.info(f"제목: {driver.title}")
        logger.info(f"열린 창: {len(driver.window_handles)}")
        logger.info(f"페이지 소스 크기: {len(driver.page_source)} bytes")
        
        # iframe 정보
        driver.switch_to.default_content()
        iframes = driver.find_elements(By.TAG_NAME, 'iframe')
        logger.info(f"iframe: {len(iframes)}개")
        
        # 주요 요소 개수
        buttons = driver.find_elements(By.TAG_NAME, 'button')
        links = driver.find_elements(By.TAG_NAME, 'a')
        inputs = driver.find_elements(By.TAG_NAME, 'input')
        logger.info(f"버튼: {len(buttons)}개, 링크: {len(links)}개, 입력필드: {len(inputs)}개")
        
    except Exception as e:
        logger.error(f"❌ 페이지 정보 조회 실패: {e}")
    logger.info("=" * 70 + "\n")


# iframe 디버그 함수: 현재 iframe 정보 및 페이지의 모든 iframe 목록 출력
def debug_iframe_info(driver, label="", show_previous=True):
    """Element 찾기 실패 또는 Timeout 발생 시 호출되어 현재 iframe 상태와 페이지의 모든 iframe을 표시합니다."""
    try:
        logger.error(f"[IFRAME DEBUG] {label}")
        
        # 이전/현재 iframe 비교
        if show_previous:
            prev_iframe = getattr(driver, '_previous_iframe', None)
            curr_iframe = getattr(driver, '_current_iframe', None)
            depth = getattr(driver, '_frame_depth', 0)
            logger.error(f"  └─ 이전 iframe: {prev_iframe or '(없음)'}")
            logger.error(f"  └─ 현재 iframe: {curr_iframe or '(루트)'}")
            logger.error(f"  └─ 프레임 깊이: {depth}")
        
        
        # 페이지의 모든 iframe 목록 (루트에서만 가능)
        try:
            driver.switch_to.default_content()
            all_iframes = driver.find_elements(By.TAG_NAME, 'iframe')
            logger.error(f"  └─ 페이지 전체 iframe 개수: {len(all_iframes)}")
            
            for idx, iframe in enumerate(all_iframes, 1):
                iframe_id = iframe.get_attribute('id') or '(id 없음)'
                iframe_name = iframe.get_attribute('name') or '(name 없음)'
                iframe_src = iframe.get_attribute('src') or '(src 없음)'
                iframe_class = iframe.get_attribute('class') or '(class 없음)'
                logger.error(f"     [{idx}] id={iframe_id} | name={iframe_name}")
                logger.error(f"         src={iframe_src[:80]}")
                logger.error(f"         class={iframe_class}")
            
            # 원래 깊이로 복귀
            for _ in range(depth):
                try:
                    driver.switch_to.frame(driver.find_elements(By.TAG_NAME, 'iframe')[0])
                except:
                    break
        except Exception as e:
            logger.error(f"  └─ iframe 목록 조회 실패: {e}")
    except Exception as e:
        logger.error(f"  └─ iframe 디버그 오류: {e}")


# ============================================================================
# 이미지 기반 클릭 함수
# ============================================================================
def locate_and_click_image(image_path, confidence=0.8, offset_x=0, offset_y=0):
    """이미지 파일을 화면에서 찾아서 클릭합니다.
    
    Args:
        image_path (str): 찾을 이미지 파일 경로
        confidence (float): 이미지 매칭 신뢰도 (0.0~1.0, 기본 0.8)
        offset_x (int): 클릭 위치 X 오프셋
        offset_y (int): 클릭 위치 Y 오프셋
    
    Returns:
        dict: {'found': bool, 'position': (x, y), 'message': str}
    """
    if not PYAUTO_AVAILABLE:
        logger.error("❌ pyautogui가 설치되지 않았습니다. pip install pyautogui")
        return {'found': False, 'message': 'pyautogui not available'}
    
    if not Path(image_path).exists():
        logger.error(f"❌ 이미지 파일을 찾을 수 없습니다: {image_path}")
        return {'found': False, 'message': f'Image not found: {image_path}'}
    
    try:
        logger.info(f"🔍 이미지 검색 중: {Path(image_path).name} (신뢰도: {confidence})")
        
        # OpenCV를 사용하여 고급 이미지 매칭
        if PYAUTO_OPENCV:
            import numpy as np
            
            # 스크린샷 캡처
            screenshot = pyautogui.screenshot()
            screenshot_array = np.array(screenshot)
            
            # 이미지 로드 (한글 경로 지원)
            # OpenCV는 한글 경로를 지원하지 않으므로 numpy로 우회
            try:
                # 파일 존재 확인
                if not Path(image_path).exists():
                    logger.error(f"❌ 이미지 파일이 존재하지 않습니다: {image_path}")
                    return {'found': False, 'message': f'Image file not found: {image_path}'}
                
                # 한글 경로 처리: numpy로 읽고 OpenCV 형식으로 변환
                with open(image_path, 'rb') as f:
                    image_bytes = np.frombuffer(f.read(), dtype=np.uint8)
                    template = cv2.imdecode(image_bytes, cv2.IMREAD_COLOR)
                
                if template is None:
                    logger.error(f"❌ 이미지 디코딩 실패: {image_path}")
                    return {'found': False, 'message': 'Failed to decode image'}
            except Exception as e:
                logger.error(f"❌ 이미지 로드 오류: {e}")
                return {'found': False, 'message': f'Failed to load image: {e}'}
            
            # BGR을 RGB로 변환
            screenshot_cv = cv2.cvtColor(screenshot_array, cv2.COLOR_RGB2BGR)
            
            # 템플릿 매칭
            result = cv2.matchTemplate(screenshot_cv, template, cv2.TM_CCOEFF_NORMED)
            min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
            
            # 신뢰도 확인
            if max_val < confidence:
                logger.warning(f"⚠️ 이미지를 찾지 못했습니다 (신뢰도: {max_val:.2f} < {confidence})")
                return {'found': False, 'message': f'Image not found (confidence: {max_val:.2f})'}
            
            # 이미지 중심 좌표 계산
            x = max_loc[0] + template.shape[1] // 2 + offset_x
            y = max_loc[1] + template.shape[0] // 2 + offset_y
            
            logger.info(f"✓ 이미지 발견! 위치: ({x}, {y}), 신뢰도: {max_val:.2f}")
        else:
            # pyautogui 기본 이미지 로케이팅 (신뢰도 옵션 미지원)
            logger.info(f"💡 OpenCV 미설치 - 기본 이미지 로케이팅 사용")
            location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            
            if location is None:
                logger.warning(f"⚠️ 이미지를 찾을 수 없습니다")
                return {'found': False, 'message': 'Image not found'}
            
            x, y = pyautogui.center(location)
            x += offset_x
            y += offset_y
            
            logger.info(f"✓ 이미지 발견! 위치: ({x}, {y})")
        
        # 클릭 수행
        pyautogui.click(x, y, duration=0.5)
        logger.info(f"✓ 클릭 완료: ({x}, {y})")
        
        return {'found': True, 'position': (x, y), 'message': f'Image clicked at ({x}, {y})'}
    
    except Exception as e:
        logger.error(f"❌ 이미지 클릭 중 오류: {e}")
        return {'found': False, 'message': str(e)}


def locate_image(image_path, confidence=0.8):
    """이미지를 화면에서 찾아서 위치를 반환합니다.
    
    Args:
        image_path (str): 찾을 이미지 파일 경로
        confidence (float): 이미지 매칭 신뢰도 (0.0~1.0)
    
    Returns:
        tuple: (x, y) 또는 None
    """
    if not PYAUTO_AVAILABLE:
        logger.error("❌ pyautogui가 설치되지 않았습니다")
        return None
    
    if not Path(image_path).exists():
        logger.error(f"❌ 이미지 파일을 찾을 수 없습니다: {image_path}")
        return None
    
    try:
        if PYAUTO_OPENCV:
            import numpy as np
            
            screenshot = pyautogui.screenshot()
            screenshot_array = np.array(screenshot)
            
            # 한글 경로 지원: numpy로 읽고 OpenCV 형식으로 변환
            with open(image_path, 'rb') as f:
                image_bytes = np.frombuffer(f.read(), dtype=np.uint8)
                template = cv2.imdecode(image_bytes, cv2.IMREAD_COLOR)
            
            if template is None:
                return None
            
            screenshot_cv = cv2.cvtColor(screenshot_array, cv2.COLOR_RGB2BGR)
            result = cv2.matchTemplate(screenshot_cv, template, cv2.TM_CCOEFF_NORMED)
            min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
            
            if max_val < confidence:
                return None
            
            x = max_loc[0] + template.shape[1] // 2
            if template is None:
                return None
            
            screenshot_cv = cv2.cvtColor(screenshot_array, cv2.COLOR_RGB2BGR)
            result = cv2.matchTemplate(screenshot_cv, template, cv2.TM_CCOEFF)
            min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
            
            if max_val < 0:
                return None
            
            x = max_loc[0] + template.shape[1] // 2
            y = max_loc[1] + template.shape[0] // 2
            return (x, y)
        else:
            location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if location is None:
                return None
            return pyautogui.center(location)
    except Exception as e:
        logger.error(f"❌ 이미지 로케이팅 중 오류: {e}")
        return None


# ============================================================================
# Windows Hotkey 입력 함수
# ============================================================================
def press_hotkey(hotkey_str):
    """Windows hotkey를 입력합니다.
    
    Args:
        hotkey_str (str): 핫키 문자열
            형식: "ctrl+c", "alt+tab", "shift+f5", "enter", "delete" 등
            또는: "ctrl,c", "alt,tab" (쉼표 구분)
    
    Returns:
        dict: {'success': bool, 'message': str}
    """
    if not PYAUTO_AVAILABLE:
        logger.error("❌ pyautogui가 설치되지 않았습니다")
        return {'success': False, 'message': 'pyautogui not available'}
    
    if not hotkey_str:
        logger.error("❌ 핫키 문자열이 비어있습니다")
        return {'success': False, 'message': 'Empty hotkey string'}
    
    try:
        # 정규화: 대소문자 통일, 공백 제거
        hotkey_str = hotkey_str.strip().lower()
        
        # 지원하는 키 이름 매핑 (pyautogui 형식)
        key_map = {
            'ctrl': 'ctrl',
            'control': 'ctrl',
            'shift': 'shift',
            'alt': 'alt',
            'win': 'win',
            'windows': 'win',
            'tab': 'tab',
            'enter': 'enter',
            'return': 'enter',
            'space': 'space',
            'delete': 'delete',
            'del': 'delete',
            'backspace': 'backspace',
            'esc': 'esc',
            'escape': 'esc',
            'up': 'up',
            'down': 'down',
            'left': 'left',
            'right': 'right',
            'home': 'home',
            'end': 'end',
            'pageup': 'pageup',
            'pagedown': 'pagedown',
            'f1': 'f1', 'f2': 'f2', 'f3': 'f3', 'f4': 'f4',
            'f5': 'f5', 'f6': 'f6', 'f7': 'f7', 'f8': 'f8',
            'f9': 'f9', 'f10': 'f10', 'f11': 'f11', 'f12': 'f12',
        }
        
        # '+' 또는 ','로 구분된 키 조합 파싱
        if '+' in hotkey_str:
            keys = [k.strip() for k in hotkey_str.split('+')]
        elif ',' in hotkey_str:
            keys = [k.strip() for k in hotkey_str.split(',')]
        else:
            # 단일 키
            keys = [hotkey_str]
        
        # 키 이름 검증 및 변환
        mapped_keys = []
        for key in keys:
            if key in key_map:
                mapped_keys.append(key_map[key])
            elif len(key) == 1 and key.isalnum():
                # 단일 문자 (a-z, 0-9)
                mapped_keys.append(key)
            else:
                logger.warning(f"⚠️ 지원하지 않는 키: {key}")
                mapped_keys.append(key)
        
        # hotkey 실행
        logger.info(f"⌨️ 핫키 입력: {' + '.join(mapped_keys)}")
        
        if len(mapped_keys) == 1:
            # 단일 키 입력
            pyautogui.press(mapped_keys[0])
        else:
            # 조합 키 입력 (Ctrl+C, Alt+Tab 등)
            pyautogui.hotkey(*mapped_keys)
        
        time.sleep(0.3)  # 입력 후 약간 대기
        logger.info(f"✓ 핫키 입력 완료")
        return {'success': True, 'message': f'Hotkey pressed: {hotkey_str}'}
    
    except Exception as e:
        logger.error(f"❌ 핫키 입력 중 오류: {e}")
        return {'success': False, 'message': str(e)}


def type_text(text_str, interval=0.05):
    """텍스트를 입력합니다 (Windows 파일 대화상자용).
    
    Args:
        text_str (str): 입력할 텍스트
        interval (float): 문자 사이 간격 (초)
    
    Returns:
        dict: {'success': bool, 'message': str}
    """
    if not PYAUTO_AVAILABLE:
        logger.error("❌ pyautogui가 설치되지 않았습니다")
        return {'success': False, 'message': 'pyautogui not available'}
    
    try:
        logger.info(f"⌨️ 텍스트 입력: {text_str[:50]}{'...' if len(text_str) > 50 else ''}")
        pyautogui.typewrite(text_str, interval=interval)
        time.sleep(0.3)
        logger.info(f"✓ 텍스트 입력 완료")
        return {'success': True, 'message': f'Text typed: {text_str[:50]}'}
    except Exception as e:
        logger.error(f"❌ 텍스트 입력 중 오류: {e}")
        return {'success': False, 'message': str(e)}


def copy_text(text_str):
    """클립보드를 통해 텍스트를 붙여넣습니다 (한글 지원).
    
    Args:
        text_str (str): 입력할 텍스트 (한글 포함)
    
    Returns:
        dict: {'success': bool, 'message': str}
    """
    if not PYAUTO_AVAILABLE:
        logger.error("❌ pyautogui가 설치되지 않았습니다")
        return {'success': False, 'message': 'pyautogui not available'}
    
    if not PYPERCLIP_AVAILABLE:
        logger.error("❌ pyperclip이 설치되지 않았습니다")
        return {'success': False, 'message': 'pyperclip not available'}
    
    try:
        logger.info(f"📋 클립보드 붙여넣기: {text_str[:50]}{'...' if len(text_str) > 50 else ''}")
        # 클립보드에 복사
        pyperclip.copy(text_str)
        time.sleep(0.1)
        # Ctrl+V로 붙여넣기
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.3)
        logger.info(f"✓ 클립보드 붙여넣기 완료")
        return {'success': True, 'message': f'Text pasted: {text_str[:50]}'}
    except Exception as e:
        logger.error(f"❌ 클립보드 붙여넣기 중 오류: {e}")
        return {'success': False, 'message': str(e)}


# %%
# 단계 1: EXCEL 파일 읽기
def read_excel_file(excel_path):
    """Excel 파일에서 Config와 Actions 읽기"""
    
    logger.info(f"Excel 파일 읽기 시작: {excel_path}")
    
    if not Path(excel_path).exists():
        logger.error(f"❌ 파일을 찾을 수 없습니다: {excel_path}")
        return None, None
    
    try:
        wb = load_workbook(excel_path)
        
        # Config 시트 읽기
        config_data = {}
        if "Config" in wb.sheetnames:
            ws_config = wb["Config"]
            for row in ws_config.iter_rows(min_row=2, values_only=True):
                if row[0]:  # 키가 있으면
                    config_data[row[0]] = row[1]  # key: value
            logger.info(f"✓ Config 읽음: {len(config_data)}개 항목")
            logger.info(f"  {config_data}")
        
        # Actions 시트 읽기
        actions_data = []
        if "Actions" in wb.sheetnames:
            ws_actions = wb["Actions"]
            raw_headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws_actions[1]]

            def canonical_key(h):
                key = (h or '').strip().lower().replace(' ', '')
                # 번호 계열
                if key in ('번호', '순번', 'no', 'num', 'number', 'index', 'idx'):
                    return '번호'
                # 타입 계열
                if key in ('타입', 'type'):
                    return '타입'
                # 액션명/이름
                if key in ('액션명', 'name', 'action', 'actionname'):
                    return '액션명'
                # 로케이터
                if key in ('xpath', 'x_path'):
                    return 'XPath'
                if key in ('css', 'cssselector'):
                    return 'CSS'
                if key in ('id',):
                    return 'id'
                if key in ('nameattr', 'namelocator'):
                    return 'name'
                # 값/파라미터
                if key in ('값', 'value', 'val', 'param'):
                    return '값'
                # 대기시간
                if key in ('대기시간', 'delay', 'wait', 'waitsec'):
                    return '대기시간'
                return h  # 원본 유지

            # 정규화된 헤더 목록
            norm_headers = [canonical_key(h) for h in raw_headers]

            for row in ws_actions.iter_rows(min_row=2, values_only=True):
                # 빈 행은 건너뜀
                if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                    continue
                entry = dict(zip(norm_headers, row))
                # 번호를 문자열로 정규화 (하이픈/점 지원)
                if '번호' in entry and entry['번호'] is not None:
                    entry['번호'] = str(entry['번호']).strip()
                actions_data.append(entry)

            logger.info(f"✓ Actions 읽음: {len(actions_data)}개 액션")
            for i, action in enumerate(actions_data, 1):
                logger.info(f"  {i}. {action}")
        
        wb.close()
        return config_data, actions_data
        
    except Exception as e:
        logger.error(f"❌ Excel 읽기 오류: {e}")
        return None, None


# %% 단계 2: 브라우저 초기화 및 사이트 접속# %% 
def init_browser(headless=False):
    """Chrome 브라우저 초기화
    
    포트 확인 후 빠르게 판단 (socket으로 1초 타임아웃).
    포트 열려있으면 기존 Chrome 연결, 없으면 새 Chrome 시작.
    """
    global config_data

    logger.info("Chrome 브라우저 초기화 중...")
    
    # 디버거 주소 설정
    debugger_addr = "127.0.0.1:9222"
    driver = None
    
    # 1. 포트가 열려있는지 빠르게 확인 (socket으로 1초 타임아웃)
    if debugger_addr:
        host, port = debugger_addr.split(':')
        port = int(port)
        
        logger.info(f"🔍 기존 Chrome 포트 확인 중: {debugger_addr}...")
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)  # 1초 타임아웃
            result = sock.connect_ex((host, port))
            sock.close()
            
            if result == 0:
                logger.info(f"✓ 포트 {port}가 열려있음. 기존 Chrome 연결 시도...")
                
                try:
                    options = webdriver.ChromeOptions()
                    if headless:
                        options.add_argument("--headless")
                    options.add_argument("--window-size=1920,1100")
                    options.add_argument("--no-sandbox")
                    options.add_argument("--disable-dev-shm-usage")
                    
                    options.add_experimental_option("debuggerAddress", debugger_addr)
                    
                    # chromedriver 생성
                    driver = webdriver.Chrome(options=options)
                    
                    # 빠른 연결 확인
                    _ = driver.current_url
                    logger.info(f"✓ debuggerAddress 연결 성공: {driver.current_url}")
                    return driver
                    
                except Exception as e:
                    logger.warning(f"⚠️ 포트는 열려있으나 Chrome 연결 실패: {e}")
                    driver = None
            else:
                logger.info(f"⚠️ 포트 {port}가 닫혀있음. 새 Chrome 시작...")
                
        except Exception as e:
            logger.warning(f"⚠️ 포트 확인 실패: {e}")

    # 2. 기존 Chrome 연결 실패 또는 포트 없음 → 새 Chrome 시작
    if driver is None:
        try:
            options = webdriver.ChromeOptions()
            if headless:
                options.add_argument("--headless")
            options.add_argument("--window-size=1920,1100")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            
            logger.info("🆕 새 Chrome 프로세스 시작 중...")
            
            try:
                driver = webdriver.Chrome(options=options)
                logger.info("✓ 새 Chrome 브라우저 시작됨")
                return driver
            except Exception as e:
                logger.warning(f"기본 chromedriver 사용: {e}")
                try:
                    # webdriver-manager로 자동 다운로드
                    from webdriver_manager.chrome import ChromeDriverManager
                    service = Service(ChromeDriverManager().install())
                    driver = webdriver.Chrome(service=service, options=options)
                    logger.info("✓ Chrome 브라우저 시작됨 (webdriver-manager)")
                    return driver
                except Exception as e2:
                    logger.error(f"❌ Chrome 초기화 실패: {e2}")
                    return None
                    
        except Exception as e:
            logger.error(f"❌ 브라우저 초기화 오류: {e}")
            return None


def access_website(driver, url):
    """웹사이트 접속"""
    
    if not driver:
        logger.error("❌ 드라이버가 초기화되지 않았습니다")
        return False
    
    logger.info(f"웹사이트 접속 중: {url}")
    
    try:
        driver.get(url)
        
        # 페이지 로드 대기 (최대 10초)
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        
        logger.info(f"✓ 접속 성공: {driver.title}")
        return True
        
    except Exception as e:
        logger.error(f"❌ 접속 실패: {e}")
        return False


# %% 파일 수집 유틸리티

def collect_files_from_config(config: dict) -> list:
    """Config의 upload_path와 common_upload_path에서 모든 파일을 수집합니다.
    
    Parameters:
    - config: Config dict (upload_path, common_upload_path 사용)
    
    Returns:
    - 파일 Path 객체 리스트 (일반경로 파일 먼저, 그 다음 공통경로 파일)
    """
    all_files = []
    
    # 일반 업로드 경로
    upload_path = config.get('upload_path') if config else None
    if upload_path:
        upath = Path(upload_path).expanduser()
        if not upath.is_absolute():
            upath = Path(__file__).parent / upath
        
        if upath.exists():
            files = sorted([f for f in upath.glob('*') if f.is_file()])
            logger.info(f"  📂 일반경로: {upath}")
            logger.info(f"     찾음: {len(files)}개")
            for idx, f in enumerate(files, 1):
                logger.info(f"       {idx}. {f.name}")
            all_files.extend(files)
        else:
            logger.warning(f"  ⚠️  일반경로 없음: {upath}")
    
    # 공통 업로드 경로
    common_upload_path = config.get('common_upload_path') if config else None
    if common_upload_path:
        cpath = Path(common_upload_path).expanduser()
        if not cpath.is_absolute():
            cpath = Path(__file__).parent / cpath
        
        if cpath.exists():
            files = sorted([f for f in cpath.glob('*') if f.is_file()])
            logger.info(f"  📂 공통경로: {cpath}")
            logger.info(f"     찾음: {len(files)}개")
            for idx, f in enumerate(files, 1):
                logger.info(f"       {idx}. {f.name}")
            all_files.extend(files)
        else:
            logger.warning(f"  ⚠️  공통경로 없음: {cpath}")
    
    if not all_files:
        logger.warning("  ⚠️  업로드할 파일이 없습니다")
    
    return all_files


def collect_upload_files(patterns: str, config: dict) -> list:
    """파일 패턴으로 업로드할 파일 목록을 수집합니다.
    
```    Parameters:
    - patterns: 파일 패턴 (;로 구분, common: 접두사로 공통 폴더 지정)
    - config: Config dict (upload_path, common_upload_path 사용)
    
    Returns:
    - 파일 Path 객체 리스트
    """
    if not patterns:
        raise ValueError('파일 패턴이 비어 있습니다')
    
    pattern_list = [p.strip() for p in patterns.split(';') if p.strip()]
    all_files = []
    
    logger.info(f"  📝 파싱된 패턴: {pattern_list}")
    
    # Config에서 경로 추출
    upload_path = config.get('upload_path') if config else None
    common_upload_path = config.get('common_upload_path') if config else None
    
    logger.info(f"  📂 일반경로: {upload_path}")
    logger.info(f"  📂 공통경로: {common_upload_path}")
    
    # 각 패턴에 해당하는 파일 수집
    for pattern in pattern_list:
        if pattern.startswith('common:'):
            # 공통 폴더에서 파일 찾기
            if not common_upload_path:
                raise ValueError('Config에 common_upload_path가 없습니다')
            
            common_path = Path(common_upload_path).expanduser()
            if not common_path.is_absolute():
                common_path = Path(__file__).parent / common_path
            
            search_str = pattern.split(':', 1)[1].strip()
            glob_pattern = f"*{search_str}*"
            
            logger.info(f"  🔍 검색중 [공통]: {common_path} / {glob_pattern}")
            
            found_files = sorted(common_path.glob(glob_pattern))
            found_files = [f for f in found_files if f.is_file()]
            
            logger.info(f"     찾음: {len(found_files)}개")
            for f in found_files:
                logger.info(f"       - {f.name}")
            
            if not found_files:
                raise ValueError(f'파일을 찾을 수 없습니다: {common_path.name}/{glob_pattern}')
            all_files.extend(found_files)
        else:
            # 일반 폴더에서 파일 찾기
            if not upload_path:
                raise ValueError('Config에 upload_path가 없습니다')
            
            upath = Path(upload_path).expanduser()
            if not upath.is_absolute():
                upath = Path(__file__).parent / upath
            
            if not upath.exists():
                raise ValueError(f'업로드 경로가 없습니다: {upath}')
            
            glob_pattern = f"*{pattern}*"
            logger.info(f"  🔍 검색중 [일반]: {upath} / {glob_pattern}")
            
            found_files = sorted(upath.glob(glob_pattern))
            found_files = [f for f in found_files if f.is_file()]
            
            logger.info(f"     찾음: {len(found_files)}개")
            for f in found_files:
                logger.info(f"       - {f.name}")
            
            if not found_files:
                raise ValueError(f'파일을 찾을 수 없습니다: {upath.name}/{glob_pattern}')
            all_files.extend(found_files)
    
    if not all_files:
        raise ValueError('업로드할 파일이 없습니다')
    
    return all_files


# %% 액션 실행 함수

def perform_ui_action(driver, action: dict, config: dict = None, test_mode: bool = False, timeout: int = 10, 
                     actions_list: list = None, current_index: int = 0):
    """단일 UI 액션을 수행합니다.
    
    Parameters:
    - timeout: WebDriverWait 대기시간 (초) - Config의 global_wait_time에서 전달됨
    - actions_list: 전체 액션 리스트 (loop_click에서 다음 액션들을 읽기 위해 필요)
    - current_index: 현재 액션의 인덱스 (0-based)

    지원 타입:
    - navigate | 페이지 이동: URL로 이동 (값 사용)
    - click   | 클릭: 요소 클릭
    - input   | 입력: 텍스트 입력 (값, 플레이스홀더 치환 지원)
    - send_keys: 특수 키 입력 (예: {ENTER}, {TAB})
    - read    | get_text | 읽기: 요소의 텍스트 반환
    - hover   | mouseover | 마우스 오버: 요소에 마우스 오버
    - select  | 드롭다운: Select 요소에서 값/텍스트 선택 (값: 'text:xxx' 또는 'value:yyy')
    - wait    | 대기: 지정 초만큼 대기 (값이 초)
    - wait_element | 요소대기: 요소가 보일 때까지 대기
    - scroll  | 스크롤: 요소로 스크롤 또는 페이지 스크롤 (값: 'element' 또는 'page:down')
    - image_click | image: 화면에 보이는 캡처 이미지 매칭 후 클릭 (값: 이미지 파일 경로)
    - iframe_in: 지정된 iframe으로 전환 (XPath/CSS/id/name 또는 값: 'index:N' | 'name:XXX' | 'id:YYY')
    - iframe_out: 프레임 복귀 (값: 'parent' | 'default' | 'levels:N')
    - upload | 파일업로드: 파일 프리픽스로 매칭된 모든 파일을 하나씩 업로드 (값: 파일 프리픽스, Config의 upload_path 사용)
    - clipboard_upload | 클립보드업로드: 파일을 클립보드에 복사 후 Ctrl+V로 업로드 (값: 파일 프리픽스;common:파일명, Config의 upload_path/common_upload_path 사용)

    action dict 예상 키:
    - '타입' (type)
    - 'XPath' / 'CSS' / 'id' / 'name'
    - '값' (value)
    - '대기시간' (optional, seconds)
    """
    # 내부 유틸 (외부에 노출하지 않음)
    def resolve(val):
        if not isinstance(val, str):
            return val
        if config:
            for k, v in config.items():
                ph = '{' + str(k) + '}'
                if ph in val:
                    val = val.replace(ph, str(v))
        return val

    def locate(wait_timeout=None):
        if action.get('XPath'):
            by, locator = By.XPATH, action['XPath']
        elif action.get('CSS'):
            by, locator = By.CSS_SELECTOR, action['CSS']
        elif action.get('id'):
            by, locator = By.ID, action['id']
        elif action.get('name'):
            by, locator = By.NAME, action['name']
        else:
            raise ValueError('Locator(XPath/CSS/id/name)가 없습니다')
        
        logger.debug(f"  [LOCATE] 요소 찾기 시작: {by} = {locator}")
        
        try:
            element = WebDriverWait(driver, wait_timeout or timeout).until(EC.presence_of_element_located((by, locator)))
            logger.debug(f"  [LOCATE] ✓ 요소 발견 성공")
            return element
        except TimeoutException as e:
            logger.error(f"  [LOCATE] ❌ Timeout: {locator}")
            
            # 현재 창 정보 출력
            try:
                logger.error(f"  ❌ 현재 URL: {driver.current_url}")
                logger.error(f"  ❌ 현재 창: {driver.current_window_handle[:20]}...")
                logger.error(f"  ❌ 전체 창 개수: {len(driver.window_handles)}")
                
                # 모든 창의 URL 출력
                for i, handle in enumerate(driver.window_handles):
                    try:
                        driver.switch_to.window(handle)
                        logger.error(f"    [창 {i}] {handle[:20]}... → {driver.current_url[:60]}...")
                    except:
                        pass
                
                # 원래 창으로 복구
                driver.switch_to.window(driver.window_handles[-1])
            except Exception as debug_e:
                logger.warning(f"⚠️ 창 정보 출력 실패: {debug_e}")
            
            # Element 찾기 timeout 발생 시 진단 정보 출력
            diagnose_element_issue(driver, locator)
            debug_iframe_info(driver, label=f"Element '{locator}' not found (timeout)", show_previous=True)
            
            # 공통 에러 처리 함수 호출 (재시도 함수: 다시 요소 찾기)
            def retry_locate():
                return WebDriverWait(driver, wait_timeout or timeout).until(EC.presence_of_element_located((by, locator)))
            
            return handle_action_error(e, 'locate', locator, retry_locate)

    result = {'status': 'ok', 'value': None, 'message': ''}
    try:
        a_type = (action.get('타입') or action.get('type') or '').strip().lower()
        
        # 액션 데이터 디버깅 - 0도 유효한 값으로 처리
        raw_value = action.get('값')
        if raw_value is None:
            raw_value = action.get('value')
        logger.debug(f"  [DEBUG] raw_value: {repr(raw_value)}, type: {type(raw_value)}")
        
        value = resolve(raw_value)
        logger.debug(f"  [DEBUG] resolved_value: {repr(value)}, type: {type(value)}")
        
        delay = action.get('대기시간') or action.get('delay') or 0
        try:
            delay = float(delay) if delay else 0
        except Exception:
            delay = 0

        # TEST MODE: 로그만 출력하고 실제 동작은 스킵 (단, loop_click/업로드 류는 내부 시뮬레이션 로직 실행)
        passthrough_types = (
            'loop_click', '루프클릭', 'repeat_click',
            'upload', '파일업로드',
            'clipboard_upload', '클립보드업로드'
        )
        if test_mode and a_type not in passthrough_types:
            if a_type in ('type_text', '텍스트입력', 'keyboard_input'):
                planned_text = str(value) if value is not None else '(빈 값)'
                result['message'] = f"[TEST] type_text 예정 입력: {planned_text[:50]}"
                logger.info(f"  🧪 {result['message']}")
                return result
            result['message'] = f'[TEST] {a_type}: {value or action.get("XPath") or action.get("id") or "no-locator"}'
            logger.info(f"  🧪 {result['message']}")
            return result

        if a_type in ('navigate', '페이지 이동'):
            if not value:
                raise ValueError('navigate에 URL 값이 필요합니다')
            driver.get(value)
            WebDriverWait(driver, timeout).until(lambda d: d.execute_script("return document.readyState") == "complete")
            result['message'] = f'Navigated to {value}'

        elif a_type in ('switch_site', '사이트이동'):
            if not value:
                raise ValueError('switch_site에 URL 값이 필요합니다')
            logger.info(f"  새 탭으로 이동: {value}")
            driver.switch_to.new_window('tab')
            driver.get(value)
            # 페이지 로드 완료 대기
            WebDriverWait(driver, timeout).until(lambda d: d.execute_script("return document.readyState") == "complete")
            
            # JavaScript 초기화 대기 (중요!)
            time.sleep(2)  # 기본 2초 대기
            if delay > 0:
                logger.info(f"  추가 대기: {delay}초")
                time.sleep(delay)
            
            result['message'] = f'Switched to new tab: {value}'

        elif a_type in ('switch_window', '창전환', 'window_switch'):
            """다른 창/탭으로 전환합니다 (iframe과 다름)."""
            # 값이 없으면 최신 창으로 자동 전환
            if not value:
                logger.info(f"  최신 창으로 자동 전환")
                switch_to_latest_window(driver)
                result['message'] = 'Switched to latest window'
            else:
                # 값이 숫자면 창 인덱스, 문자열이면 URL 패턴 매칭
                try:
                    window_index = int(value)
                    handles = driver.window_handles
                    
                    if window_index < 0 or window_index >= len(handles):
                        raise ValueError(f'창 인덱스 범위 초과: {window_index} (총 {len(handles)}개)')
                    
                    target_handle = handles[window_index]
                    driver.switch_to.window(target_handle)
                    logger.info(f"  창 전환: 인덱스 {window_index} → {driver.current_url}")
                    
                    # 페이지 로드 대기
                    time.sleep(1)
                    try:
                        WebDriverWait(driver, 5).until(
                            lambda d: d.execute_script("return document.readyState") == "complete"
                        )
                    except:
                        pass
                    
                    result['message'] = f'Switched to window {window_index}: {driver.current_url}'
                
                except ValueError as e:
                    # 숫자가 아니면 URL 패턴으로 검색
                    pattern = str(value).lower()
                    handles = driver.window_handles
                    found = False
                    
                    for i, handle in enumerate(handles):
                        driver.switch_to.window(handle)
                        current_url = driver.current_url.lower()
                        
                        if pattern in current_url:
                            logger.info(f"  URL 패턴 매칭 성공: '{pattern}' in {driver.current_url}")
                            logger.info(f"  창 전환: 인덱스 {i}")
                            found = True
                            
                            # 페이지 로드 대기
                            time.sleep(1)
                            try:
                                WebDriverWait(driver, 5).until(
                                    lambda d: d.execute_script("return document.readyState") == "complete"
                                )
                            except:
                                pass
                            
                            result['message'] = f'Switched to window with URL pattern: {pattern}'
                            break
                    
                    if not found:
                        raise ValueError(f"URL 패턴 '{pattern}'과 일치하는 창을 찾을 수 없습니다")

        elif a_type in ('click', '클릭'):
            elem = locate()
            element_info = action.get('XPath') or action.get('CSS') or action.get('id') or action.get('name')
            
            # 클릭 가능 대기 (가능하면 XPath 기준으로)
            click_cond = EC.element_to_be_clickable((By.XPATH, action.get('XPath'))) if action.get('XPath') else EC.visibility_of(elem)
            WebDriverWait(driver, timeout).until(click_cond)
            
            # 클릭 시도 with 공통 에러 처리
            try:
                # 1. 요소를 화면 중앙으로 스크롤
                driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'instant'});", elem)
                time.sleep(0.3)  # 스크롤 완료 대기
                
                # 2. 일반 클릭 시도
                elem.click()
            except ElementClickInterceptedException as e:
                logger.warning(f"⚠️ 일반 클릭 차단됨, JavaScript 클릭 시도...")
                
                # JavaScript 클릭으로 재시도하는 함수
                def retry_click():
                    # 오버레이/팝업 제거 시도
                    try:
                        driver.execute_script("""
                            // 고정 요소 중 클릭을 가로막을 수 있는 것들 숨기기
                            document.querySelectorAll('.modal-backdrop, .overlay, [class*="popup"]').forEach(el => {
                                if (getComputedStyle(el).position === 'fixed' || getComputedStyle(el).position === 'absolute') {
                                    el.style.display = 'none';
                                }
                            });
                        """)
                    except:
                        pass
                    
                    # JavaScript 클릭
                    driver.execute_script("arguments[0].click();", elem)
                    logger.info("✓ JavaScript 클릭 성공")
                
                handle_action_error(e, 'click', element_info, retry_click)
            
            result['message'] = 'Clicked element'

        elif a_type in ('xpath_click', 'xp_click', 'xpath클릭'):
            """XPath만 사용하여 간단하게 클릭합니다."""
            if not action.get('XPath'):
                raise ValueError('xpath_click에 XPath가 필요합니다')
            
            xpath = action.get('XPath')
            logger.info(f"  XPath 직접 클릭: {xpath}")
            
            try:
                elem = locate()
                # 스크롤 + 클릭
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elem)
                time.sleep(0.2)
                elem.click()
                logger.info("✓ XPath 클릭 성공")
                result['message'] = 'Clicked by XPath'
            except Exception as e:
                logger.error(f"❌ XPath 클릭 실패: {e}")
                # JavaScript 클릭으로 재시도
                try:
                    driver.execute_script("arguments[0].click();", elem)
                    logger.info("✓ JavaScript로 XPath 클릭 성공")
                    result['message'] = 'Clicked by XPath (JavaScript)'
                except Exception as e2:
                    logger.error(f"❌ XPath 클릭 완전 실패: {e2}")
                    raise

        elif a_type in ('image_click', '이미지클릭'):
            """이미지 파일을 찾아서 클릭합니다."""
            if not value:
                raise ValueError('image_click에 이미지 파일 경로(값)가 필요합니다')
            
            # 신뢰도 및 오프셋 파라미터
            confidence = float(action.get('신뢰도') or action.get('confidence') or 0.8)
            offset_x = int(action.get('오프셋X') or action.get('offset_x') or 0)
            offset_y = int(action.get('오프셋Y') or action.get('offset_y') or 0)
            retry_count = int(action.get('재시도') or action.get('retry') or 1)  # 기본 1회 재시도
            retry_delay = float(action.get('재시도대기') or action.get('retry_delay') or 2.0)  # 재시도 간격 2초
            
            logger.info(f"🖼️ 이미지로 클릭: {value} (신뢰도: {confidence})")
            
            # 이미지 파일 경로 해석 (상대 경로 지원)
            image_path = Path(value)
            if not image_path.is_absolute():
                # 현재 스크립트 디렉토리 기준으로 상대 경로 해석
                image_path = Path(__file__).parent / value
            
            # 재시도 로직
            result_dict = None
            for attempt in range(retry_count + 1):
                if attempt > 0:
                    logger.info(f"  재시도 {attempt}/{retry_count} (대기: {retry_delay}초)")
                    time.sleep(retry_delay)
                
                result_dict = locate_and_click_image(
                    str(image_path),
                    confidence=confidence,
                    offset_x=offset_x,
                    offset_y=offset_y
                )
                
                if result_dict['found']:
                    if attempt > 0:
                        logger.info(f"  ✓ 재시도 {attempt}회 만에 성공!")
                    result['message'] = result_dict['message']
                    break
                else:
                    if attempt < retry_count:
                        logger.warning(f"  ⚠️ 시도 {attempt + 1} 실패: {result_dict['message']}")
            
            # 모든 재시도 실패5
            if not result_dict['found']:
                raise Exception(f"이미지를 찾을 수 없음 ({retry_count + 1}회 시도): {result_dict['message']}")

        elif a_type in ('loop_click', '루프클릭', 'repeat_click'):
            """파일 개수만큼 버튼 클릭 + 서브 액션 반복 실행
            
            루프 개수 결정:
            - 값 비움 (추천): Config의 upload_path에서 실제 파일 개수 감지
            - 값 명시: 고정 루프 횟수
            
            반복할 액션 그룹:
            - loop_click 바로 다음 액션들 (점 표기법: 22.1, 22.2, ...)
            - 정수 번호가 나오면 메인 루프 복귀
            
            Excel 예시:
              22    loop_click     //*[@id="button_{i}_4"]      (빈칸)     ← 파일 개수만큼 반복, XPath에 {i} 포함
              22.1  type_text                                   (빈칸)     ← 파일 경로 자동 주입
              22.2  hotkey                                      enter     ← Enter 확인
              23    click          //*[@id="submit"]                      ← 메인 루프 복귀
            """
            if not actions_list:
                raise ValueError('loop_click은 actions_list 파라미터가 필요합니다')
            
            # XPath 템플릿 추출 ({i} 포함해야 함)
            xpath_template = action.get('XPath') or action.get('xpath')
            if not xpath_template:
                raise ValueError('loop_click에 XPath 템플릿({i} 포함)이 필요합니다')
            
            if '{i}' not in xpath_template:
                raise ValueError(f'loop_click의 XPath에 {{i}} 플레이스홀더가 없습니다: {xpath_template}')
            
            # ========== 1단계: 루프 개수 결정 (setup_config에서 수집한 파일 목록 사용) ==========
            if not value:
                # 자동 감지: setup_config()에서 이미 수집한 파일 목록 사용
                logger.info(f"  📊 루프 개수 결정 중 (setup_config에서 수집한 파일)...")
                
                # setup_config에서 저장한 파일 목록 사용
                all_files = config.get('_collected_files') if config else []
                
                if not all_files:
                    raise ValueError('setup_config에서 파일을 수집하지 못했습니다. 업로드 폴더를 확인하세요.')
                
                loop_count = len(all_files)
                logger.info(f"  ✓ 루프 개수: {loop_count}개 (파일: {[f.name for f in all_files]})")
            else:
                # 명시된 값 사용
                try:
                    loop_count = int(value)
                    logger.info(f"  ✓ 명시된 루프: {loop_count}회")
                except:
                    raise ValueError(f'loop_click 값은 정수여야 합니다: {value}')
            
            # ========== 2단계: 반복할 액션 그룹 수집 ==========
            # Excel 번호를 기준으로 그룹 판정: 
            # - 22번이 loop_click이면 22.1, 22.2, 22.3... 형식의 액션들이 반복 그룹
            # - 23번 (정수)이 나오면 메인 루프 복귀
            
            current_num = action.get('번호') or action.get('순번') or action.get('number') or str(current_index + 1)
            current_num_str = str(current_num).strip()
            
            logger.info(f"  📦 현재 번호: {current_num_str}, 반복 그룹 수집 중...")
            
            loop_actions = []
            for j in range(current_index + 1, len(actions_list)):
                next_action = actions_list[j]
                next_num = next_action.get('번호') or next_action.get('순번') or next_action.get('number') or str(j + 1)
                next_num_str = str(next_num).strip()
                
                
                # 번호가 "22.1", "22.2" 또는 "22-1", "22-2" 형식인지 확인 (점 또는 하이픈이 있으면 서브 액션)
                if '.' in next_num_str or '-' in next_num_str:
                    # 점 또는 하이픈 앞 부분이 현재 번호와 같으면 반복 그룹
                    separator = '.' if '.' in next_num_str else '-'
                    prefix = next_num_str.split(separator)[0]
                    if prefix == current_num_str:
                        loop_actions.append(next_action)
                        logger.info(f"     - [{next_num_str}] {next_action.get('액션명') or next_action.get('name')} (반복 그룹 포함)")
                    else:
                        break  # 다른 그룹의 서브 액션
                else:
                    # 정수 번호면 메인 액션이므로 종료
                    break
                
                # 최대 20개까지만 (무한 루프 방지)
                if len(loop_actions) >= 20:
                    logger.warning(f"  ⚠️ 반복 그룹이 20개를 초과하여 중단합니다.")
                    break

            # 하이픈/점 번호 기반 서브 액션만 포함 (사용자 요구 사항)
            
            logger.info(f"  📦 반복 그룹: {len(loop_actions)}개 액션 감지")
            for idx_sub, sub_act in enumerate(loop_actions, 1):
                sub_name = sub_act.get('액션명') or sub_act.get('name') or f'Sub-{idx_sub}'
                logger.info(f"     {idx_sub}. {sub_name}")
            if test_mode and len(loop_actions) == 0:
                logger.info("  🧪 [TEST] 서브 액션이 없습니다 (type_text가 없으면 파일 경로 입력도 생략됩니다)")
            
            # ========== 3단계: 루프 실행 ==========
            logger.info(f"  🔄 루프 시작: {loop_count}회 반복")
            logger.info(f"     - 각 루프마다: 버튼 클릭 → 서브 액션 {len(loop_actions)}개 실행")
            
            success_count = 0
            
            # 수집된 파일 목록 (setup_config에서 저장함)
            collected_files = config.get('_collected_files', []) if config else []
            
            for i in range(loop_count):
                # 현재 파일 경로
                current_file_path = collected_files[i] if i < len(collected_files) else None
                logger.info(f"\n  ▶ [{i+1}/{loop_count}] 반복 시작 (파일: {current_file_path.name if current_file_path else 'N/A'})")
                if test_mode and current_file_path:
                    logger.info(f"    🧪 [TEST] 이번 루프 파일 절대경로: {current_file_path.absolute()}")
                
                # ========== 3-1: 버튼 클릭 (XPath의 {i}를 현재 인덱스로 치환) ==========
                current_xpath = xpath_template.replace('{i}', str(i))
                logger.info(f"    🖱️  {i}번 버튼 클릭 - 최종 XPath: {current_xpath}")

                if test_mode:
                    logger.info(f"      🧪 [TEST] 버튼 클릭 시뮬레이션: {current_xpath}")
                    time.sleep(0.1)
                else:
                    try:
                        elem = driver.find_element(By.XPATH, current_xpath)
                        
                        # 스크롤 + 클릭
                        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elem)
                        time.sleep(0.5)
                        
                        try:
                            elem.click()
                            logger.info(f"      ✓ 클릭 성공")
                        except ElementClickInterceptedException:
                            # 오버레이 제거 후 재시도
                            logger.warning(f"      ⚠️ ElementClickIntercepted, 오버레이 제거")
                            driver.execute_script("""
                                let overlays = document.querySelectorAll('[style*="position: fixed"], [style*="position: absolute"]');
                                overlays.forEach(o => o.style.display = 'none');
                            """)
                            time.sleep(0.3)
                            
                            try:
                                elem.click()
                                logger.info(f"      ✓ 클릭 성공 (오버레이 제거 후)")
                            except:
                                # 자바스크립트 클릭
                                driver.execute_script("arguments[0].click();", elem)
                                logger.info(f"      ✓ 클릭 성공 (자바스크립트)")
                        
                        # 클릭 후 대기 (파일 다이얼로그 열릴 시간)
                        time.sleep(1)
                        
                    except NoSuchElementException:
                        logger.error(f"      ✗ 버튼을 찾을 수 없음: {current_xpath}")
                        continue
                    except Exception as e:
                        logger.error(f"      ✗ 버튼 클릭 실패: {type(e).__name__}: {str(e)[:100]}")
                        continue
                
                # ========== 3-2: 서브 액션 그룹 실행 (재귀 호출 없이 직접 처리) ==========
                for sub_idx, sub_action in enumerate(loop_actions, 1):
                    # {i}를 현재 인덱스로 치환
                    sub_action_resolved = {}
                    for k, v in sub_action.items():
                        if isinstance(v, str):
                            sub_action_resolved[k] = v.replace('{i}', str(i))
                        else:
                            sub_action_resolved[k] = v
                    
                    sub_name = sub_action_resolved.get('액션명') or sub_action_resolved.get('name') or f'Sub-{sub_idx}'
                    sub_type = (sub_action_resolved.get('타입') or sub_action_resolved.get('type') or '').lower()
                    sub_value = resolve(sub_action_resolved.get('값') or sub_action_resolved.get('value'))
                    
                    # {current_file} 플레이스홀더를 현재 파일 절대경로로 치환
                    if sub_value and '{current_file}' in str(sub_value) and current_file_path:
                        sub_value = str(sub_value).replace('{current_file}', str(current_file_path.absolute()))
                    
                    logger.info(f"    [{i+1}-{sub_idx}] {sub_name} (타입: {sub_type})")
                    
                    # ========== 서브 액션 타입별 처리 (재귀 없음) ==========
                    try:
                        # TEST MODE: 실제 동작 없이 로그만 출력
                        if test_mode:
                            if sub_type in ('hotkey', '핫키', 'windows_key', '윈도우키'):
                                logger.info(f"      🧪 [TEST] 핫키 입력 예정: {sub_value}")
                                continue
                            elif sub_type in ('type_text', '텍스트입력', 'keyboard_input'):
                                # sub_value는 이미 {current_file} 치환됨
                                logger.info(f"      🧪 [TEST] 텍스트 입력 예정: {sub_value}")
                                continue
                            elif sub_type in ('copy_text', '복사붙여넣기', 'paste_text', 'clipboard_input'):
                                logger.info(f"      🧪 [TEST] 클립보드 붙여넣기 예정: {sub_value}")
                                continue
                            elif sub_type in ('wait', '대기'):
                                wait_sec = float(sub_value) if sub_value else 1
                                logger.info(f"      🧪 [TEST] 대기 예정: {wait_sec}초")
                                continue
                            elif sub_type in ('click', '클릭'):
                                sub_xpath = sub_action_resolved.get('XPath') or sub_action_resolved.get('xpath')
                                logger.info(f"      🧪 [TEST] 클릭 예정: {sub_xpath}")
                                continue
                        
                        if sub_type in ('hotkey', '핫키', 'windows_key', '윈도우키'):
                            # Hotkey 처리
                            if not sub_value:
                                raise ValueError('hotkey에 핫키 문자열(값)이 필요합니다')
                            logger.info(f"      Windows 핫키 입력: {sub_value}")
                            hotkey_result = press_hotkey(str(sub_value))
                            if hotkey_result['success']:
                                logger.info(f"      ✓ {hotkey_result['message']}")
                            else:
                                logger.warning(f"      ✗ {hotkey_result['message']}")
                        
                        elif sub_type in ('type_text', '텍스트입력', 'keyboard_input'):
                            # 텍스트 입력 처리 ({current_file}은 이미 치환됨)
                            if not sub_value:
                                logger.warning(f"      ⚠️ 입력할 텍스트 없음, type_text 스킵")
                                continue
                            logger.info(f"      텍스트 입력: {str(sub_value)[:80]}")
                            text_result = type_text(str(sub_value))
                            if text_result['success']:
                                logger.info(f"      ✓ {text_result['message']}")
                            else:
                                logger.warning(f"      ✗ {text_result['message']}")
                        
                        elif sub_type in ('copy_text', '복사붙여넣기', 'paste_text', 'clipboard_input'):
                            # 클립보드 붙여넣기 처리 (한글 지원)
                            if not sub_value:
                                logger.warning(f"      ⚠️ 입력할 텍스트 없음, copy_text 스킵")
                                continue
                            logger.info(f"      클립보드 붙여넣기: {str(sub_value)[:80]}")
                            copy_result = copy_text(str(sub_value))
                            if copy_result['success']:
                                logger.info(f"      ✓ {copy_result['message']}")
                            else:
                                logger.warning(f"      ✗ {copy_result['message']}")
                        
                        elif sub_type in ('wait', '대기'):
                            # 대기 처리
                            wait_sec = float(sub_value) if sub_value else 1
                            logger.info(f"      대기: {wait_sec}초")
                            time.sleep(wait_sec)
                            logger.info(f"      ✓ 대기 완료")
                        
                        elif sub_type in ('click', '클릭'):
                            # 클릭 처리 (XPath 사용)
                            sub_xpath = sub_action_resolved.get('XPath') or sub_action_resolved.get('xpath')
                            if not sub_xpath:
                                raise ValueError('click에 XPath가 필요합니다')
                            logger.info(f"      클릭: {sub_xpath}")
                            sub_elem = WebDriverWait(driver, timeout).until(
                                EC.element_to_be_clickable((By.XPATH, sub_xpath))
                            )
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", sub_elem)
                            time.sleep(0.3)
                            sub_elem.click()
                            logger.info(f"      ✓ 클릭 성공")
                        
                        else:
                            # 지원하지 않는 타입은 경고
                            logger.warning(f"      ⚠️ loop_click 서브 액션에서 지원하지 않는 타입: {sub_type}")
                            logger.warning(f"      ⚠️ 지원 타입: hotkey, type_text, wait, click")
                    
                    except Exception as sub_e:
                        logger.warning(f"      ✗ 서브 액션 실패: {type(sub_e).__name__}: {str(sub_e)[:100]}")
                
                success_count += 1
                logger.info(f"  ✓ [{i+1}/{loop_count}] 반복 완료")
            
            # ========== 4단계: 결과 반환 + skip 정보 전달 ==========
            result['message'] = f'{success_count}/{loop_count}회 반복 완료 (액션 그룹: {len(loop_actions)}개)'
            result['skip_next_actions'] = len(loop_actions)  # 메인 루프에서 건너뛸 액션 개수
            
            logger.info(f"  🎯 루프 완료: 다음 {len(loop_actions)}개 액션 건너뜀")

        elif a_type in ('hotkey', '핫키', 'windows_key', '윈도우키'):
            """Windows hotkey를 입력합니다 (파일 대화상자 처리용)."""
            if not value:
                raise ValueError('hotkey에 핫키 문자열(값)이 필요합니다')
            
            logger.info(f"  Windows 핫키 입력: {value}")
            
            result_dict = press_hotkey(str(value))
            if not result_dict['success']:
                raise Exception(f"핫키 입력 실패: {result_dict['message']}")
            
            result['message'] = result_dict['message']

        elif a_type in ('type_text', '텍스트입력', 'keyboard_input'):
            """일반 텍스트를 입력합니다 (XPath가 있으면 요소 클릭 후 기존 텍스트 삭제 후 입력)."""
            if value is None or value == '':
                raise ValueError('type_text에 텍스트(값)가 필요합니다')
            
            # XPath가 있으면 먼저 해당 요소 클릭
            xpath = action.get('XPath') or action.get('xpath')
            if xpath:
                logger.info(f"  📍 입력 필드 클릭: {xpath[:60]}")
                try:
                    elem = locate()
                    elem.click()
                    time.sleep(0.3)
                    # 기존 텍스트 전체 선택 후 삭제 (Ctrl+A)
                    logger.info(f"  🔄 기존 텍스트 선택 (Ctrl+A)")
                    pyautogui.hotkey('ctrl', 'a')
                    time.sleep(0.1)
                except Exception as e:
                    logger.warning(f"  ⚠️ 요소 클릭 실패, 그대로 진행: {e}")
            
            logger.info(f"  텍스트 입력: {str(value)[:50]}")
            
            result_dict = type_text(str(value))
            if not result_dict['success']:
                raise Exception(f"텍스트 입력 실패: {result_dict['message']}")
            
            result['message'] = result_dict['message']

        elif a_type in ('copy_text', '복사붙여넣기', 'paste_text', 'clipboard_input'):
            """클립보드를 통해 텍스트를 붙여넣습니다 (한글 지원)."""
            if not value:
                raise ValueError('copy_text에 텍스트(값)가 필요합니다')
            
            logger.info(f"  클립보드 붙여넣기: {str(value)[:50]}")
            
            result_dict = copy_text(str(value))
            if not result_dict['success']:
                raise Exception(f"클립보드 붙여넣기 실패: {result_dict['message']}")
            
            result['message'] = result_dict['message']

        elif a_type in ('input', '입력'):
            if value is None:
                value = ''
            elem = locate()
            element_info = action.get('XPath') or action.get('CSS') or action.get('id') or action.get('name')
            
            try:
                elem.clear()
                elem.send_keys(str(value))
            except (StaleElementReferenceException, ElementClickInterceptedException) as e:
                def retry_input():
                    # 요소 재탐색
                    elem_retry = locate()
                    elem_retry.clear()
                    elem_retry.send_keys(str(value))
                    logger.info(f"✓ 입력 재시도 성공: {value}")
                
                handle_action_error(e, 'input', element_info, retry_input)
            
            result['message'] = f'Input sent: {value}'

        elif a_type in ('send_keys',):
            elem = locate()
            element_info = action.get('XPath') or action.get('CSS') or action.get('id') or action.get('name')
            keys_map = {
                '{ENTER}': Keys.ENTER,
                '{TAB}': Keys.TAB,
                '{ESC}': Keys.ESCAPE,
                '{SPACE}': Keys.SPACE,
            }
            key_to_send = keys_map.get(value) if isinstance(value, str) and value in keys_map else (str(value) if value is not None else '')
            
            try:
                if isinstance(key_to_send, str):
                    elem.send_keys(key_to_send)
                else:
                    elem.send_keys(key_to_send)
            except (StaleElementReferenceException, ElementClickInterceptedException) as e:
                def retry_send_keys():
                    elem_retry = locate()
                    if isinstance(key_to_send, str):
                        elem_retry.send_keys(key_to_send)
                    else:
                        elem_retry.send_keys(key_to_send)
                    logger.info(f"✓ 키 전송 재시도 성공: {value}")
                
                handle_action_error(e, 'send_keys', element_info, retry_send_keys)
            
            result['message'] = f'Keys sent: {value}'

        elif a_type in ('read', 'get_text', '읽기'):
            elem = locate()
            text = elem.text
            result['value'] = text
            result['message'] = f'Text read: {text}'

        elif a_type in ('hover', 'mouseover', '마우스 오버'):
            elem = locate()
            ActionChains(driver).move_to_element(elem).perform()
            result['message'] = 'Hovered over element'

        elif a_type in ('select', '드롭다운'):
            elem = locate()
            sel = Select(elem)
            if isinstance(value, str) and value.startswith('text:'):
                sel.select_by_visible_text(value.split(':', 1)[1])
            elif isinstance(value, str) and value.startswith('value:'):
                sel.select_by_value(value.split(':', 1)[1])
            else:
                # 기본은 visible_text로 시도
                sel.select_by_visible_text(str(value))
            result['message'] = f'Selected: {value}'

        elif a_type in ('wait', '대기'):
            sec = float(value) if value else delay
            time.sleep(sec)
            result['message'] = f'Waited {sec} sec'

        elif a_type in ('wait_element', '요소대기'):
            custom_timeout = int(value) if value else timeout
            _ = locate(wait_timeout=custom_timeout)
            result['message'] = 'Element appeared'

        elif a_type in ('scroll', '스크롤'):
            mode = (value or '').strip().lower()
            if mode.startswith('page:'):
                if mode.endswith('down'):
                    driver.execute_script('window.scrollBy(0, window.innerHeight);')
                elif mode.endswith('up'):
                    driver.execute_script('window.scrollBy(0, -window.innerHeight);')
                else:
                    driver.execute_script('window.scrollTo(0, 0);')
            else:
                elem = locate()
                driver.execute_script('arguments[0].scrollIntoView({behavior: "smooth", block: "center"});', elem)
            result['message'] = f'Scrolled: {value or "element"}'

        elif a_type in ('iframe_in', 'frame_in', 'iframe', '프레임', '프레임진입'):
            # 프레임 전환 대상 찾기: 우선 로케이터, 없으면 값으로 index/name/id 처리
            target_desc = None
            # 이전 iframe 저장
            prev_iframe = getattr(driver, '_current_iframe', None)
            setattr(driver, '_previous_iframe', prev_iframe)
            
            if action.get('XPath') or action.get('CSS') or action.get('id') or action.get('name'):
                elem = locate()
                driver.switch_to.frame(elem)
                target_desc = action.get('XPath') or action.get('CSS') or action.get('id') or action.get('name')
            else:
                sval = (value or '').strip()
                if not sval:
                    raise ValueError('iframe_in에 대상 지정이 없습니다 (XPath/CSS/id/name 또는 값)')
                if sval.startswith('index:'):
                    idx = int(sval.split(':', 1)[1])
                    driver.switch_to.frame(idx)
                    target_desc = f'index:{idx}'
                elif sval.startswith('name:'):
                    nm = sval.split(':', 1)[1]
                    driver.switch_to.frame(nm)
                    target_desc = f'name:{nm}'
                elif sval.startswith('id:'):
                    fid = sval.split(':', 1)[1]
                    driver.switch_to.frame(fid)
                    target_desc = f'id:{fid}'
                else:
                    # 기본은 name/id로 시도
                    driver.switch_to.frame(sval)
                    target_desc = sval

            # 현재 iframe 저장
            setattr(driver, '_current_iframe', target_desc)
            # 프레임 진입 로그 및 깊이 추적
            depth = getattr(driver, '_frame_depth', 0) + 1
            setattr(driver, '_frame_depth', depth)
            logger.info(f"  📍 iframe 진입: {target_desc} (깊이={depth})")
            result['message'] = f'Entered iframe ({target_desc}), depth={depth}'

        elif a_type in ('iframe_out', 'frame_out', '프레임복귀', '프레임아웃'):
            mode = (value or '').strip().lower()
            prev_iframe = getattr(driver, '_current_iframe', None)
            setattr(driver, '_previous_iframe', prev_iframe)
            
            if mode in ('parent', 'up', '..'):
                driver.switch_to.parent_frame()
                # 깊이 감소
                depth = max(0, getattr(driver, '_frame_depth', 0) - 1)
                setattr(driver, '_frame_depth', depth)
                setattr(driver, '_current_iframe', None if depth == 0 else f'parent(depth={depth})')
                logger.info(f"  📍 parent frame 복귀 (깊이={depth})")
                result['message'] = f'Returned to parent frame, depth={depth}'
            elif mode.startswith('levels:'):
                n = int(mode.split(':', 1)[1])
                for _ in range(max(1, n)):
                    driver.switch_to.parent_frame()
                depth = max(0, getattr(driver, '_frame_depth', 0) - max(1, n))
                setattr(driver, '_frame_depth', depth)
                setattr(driver, '_current_iframe', None if depth == 0 else f'up_{n}_levels(depth={depth})')
                logger.info(f"  📍 {n}단계 상위로 복귀 (깊이={depth})")
                result['message'] = f'Returned {n} level(s) up, depth={depth}'
            else:
                driver.switch_to.default_content()
                setattr(driver, '_frame_depth', 0)
                setattr(driver, '_current_iframe', None)
                logger.info(f"  📍 루트 프레임 복귀 (깊이=0)")
                result['message'] = 'Returned to default content (root frame)'

        elif a_type in ('image_click', 'image'):
            if not PYAUTO_AVAILABLE:
                raise ValueError('pyautogui가 필요합니다. pip install pyautogui opencv-python')
            img_path = resolve(value)
            if not img_path:
                raise ValueError('이미지 경로가 비어 있습니다')
            img_path = Path(img_path).expanduser()
            if not img_path.is_absolute():
                img_path = Path(__file__).parent / img_path
            if not img_path.exists():
                raise ValueError(f'이미지 파일이 없습니다: {img_path}')

            deadline = time.time() + timeout
            location = None
            while time.time() < deadline and location is None:
                try:
                    location = pyautogui.locateOnScreen(str(img_path), confidence=0.8 if PYAUTO_OPENCV else None)
                except Exception as find_err:
                    raise ValueError(f'이미지 매칭 오류: {find_err}')
                if location is None:
                    time.sleep(0.3)
            if location is None:
                raise ValueError(f'이미지를 화면에서 찾지 못했습니다: {img_path}')
            center = pyautogui.center(location)
            pyautogui.click(center.x, center.y)
            result['message'] = f'Image clicked: {img_path}'

        elif a_type in ('upload', '파일업로드'):
            # 파일 프리픽스 받기
            file_prefix = resolve(value)
            if not file_prefix:
                raise ValueError('파일 프리픽스가 비어 있습니다')
            
            # Config에서 업로드 경로 추출
            upload_path = config.get('upload_path') if config else None
            if not upload_path:
                raise ValueError('Config에 upload_path가 없습니다')
            
            upload_path = Path(upload_path).expanduser()
            if not upload_path.is_absolute():
                upload_path = Path(__file__).parent / upload_path
            
            if not upload_path.exists():
                raise ValueError(f'업로드 경로가 없습니다: {upload_path}')
            
            # 파일 찾기 (프리픽스로 시작하는 모든 파일)
            pattern = f"{file_prefix}*"
            files = sorted(upload_path.glob(pattern))
            files = [f for f in files if f.is_file()]  # 디렉토리 제외
            
            if not files:
                raise ValueError(f'파일을 찾을 수 없습니다: {upload_path.name}/{pattern}')
            
            # TEST MODE: 찾은 파일 목록 출력
            if test_mode:
                logger.info(f"  🧪 [검출된 파일 목록]")
                for idx, fp in enumerate(files, 1):
                    logger.info(f"     {idx}. {fp.absolute()}")
                result['message'] = f'[TEST] {len(files)}개 파일 업로드 가능: {[f.name for f in files]}'
                return result
            
            # 각 파일마다 하나씩 업로드
            uploaded_names = []
            for idx, file_path in enumerate(files, 1):
                elem = locate()  # 매번 요소 다시 찾기
                elem.send_keys(str(file_path.absolute()))
                uploaded_names.append(file_path.name)
                time.sleep(0.5)
                
                if idx < len(files):
                    # 다음 파일 업로드 준비
                    time.sleep(0.3)
            
            result['message'] = f'{len(files)}개 파일 업로드: {uploaded_names}'

        elif a_type in ('clipboard_upload', '클립보드업로드'):
            if not PYPERCLIP_AVAILABLE:
                raise ValueError('pyperclip이 필요합니다. pip install pyperclip')
            
            # 파일 패턴 파싱
            logger.debug(f"    [DEBUG] action.get('값') = {repr(action.get('값'))}")
            logger.debug(f"    [DEBUG] action.get('value') = {repr(action.get('value'))}")
            logger.debug(f"    [DEBUG] value variable before clipboard_upload = {repr(value)}")
            
            patterns = resolve(value)
            logger.debug(f"    [DEBUG] patterns after resolve = {repr(patterns)}")
            
            # 파일 수집 함수 호출
            all_files = collect_upload_files(patterns, config)
            
            # TEST MODE: 업로드 루프를 파일별로 시뮬레이션하며 로그 출력
            if test_mode:
                logger.info("  🧪 [업로드 시뮬레이션 시작]")
                for idx, fp in enumerate(all_files, 1):
                    logger.info(f"  ▶ [{idx}/{len(all_files)}] 업로드 예정: {fp.name}")
                    time.sleep(0.1)
                    logger.info(f"  ✔ [{idx}/{len(all_files)}] 업로드 완료(시뮬): {fp.name}")
                result['message'] = f'[TEST] {len(all_files)}개 파일 업로드 루프 확인 완료'
                return result
            
            # 클립보드를 통해 각 파일 업로드
            uploaded_names = []
            for idx, file_path in enumerate(all_files, 1):
                logger.info(f"  ▶ [{idx}/{len(all_files)}] 업로드 시작: {file_path.name}")
                # 클립보드에 파일 경로 복사
                pyperclip.copy(str(file_path.absolute()))
                # Ctrl+V로 붙여넣기
                pyautogui.hotkey('ctrl', 'v')
                time.sleep(0.5)
                # Enter로 확인
                pyautogui.press('enter')
                uploaded_names.append(file_path.name)
                logger.info(f"  ✔ [{idx}/{len(all_files)}] 업로드 완료: {file_path.name}")
                time.sleep(1)  # 업로드 처리 대기
            
            result['message'] = f'{len(all_files)}개 파일 클립보드 업로드: {uploaded_names}'

        else:
            raise ValueError(f'지원하지 않는 액션 타입: {a_type}')

        # 액션 후 추가 대기
        if delay and a_type not in ('wait',):
            time.sleep(delay)

    except TimeoutException as timeout_err:
        result['status'] = 'error'
        result['message'] = f'Timeout: {str(timeout_err)[:100]}'
        logger.error(f"  ✗ TIMEOUT 발생! {result['message']}")
        # Timeout 발생 시 iframe 디버그 정보 출력 (locate()에서 이미 출력했을 수 있지만 추가)
        debug_iframe_info(driver, label=f"Action '{action.get('액션명')}' timeout", show_previous=True)
    except Exception as e:
        result['status'] = 'error'
        result['message'] = str(e)
    return result


# %% 액션 실행기 (Runner)

def run_actions(driver, actions_list: list, config: dict = None, test_mode: bool = False):
    """Excel Actions 시트의 모든 액션을 순차 실행합니다.
    
    Parameters:
    - driver: Selenium WebDriver 인스턴스
    - actions_list: Action dict 리스트
    - config: Config dict (플레이스홀더 치환용 + global_wait_time)
    - test_mode: True이면 로그만 출력하고 실제 동작 스킵
    
    Returns: 성공/실패 결과 리스트
    """
    # Config에서 global_wait_time 추출 (기본값: 10초)
    global_wait_time = int(config.get('global_wait_time', 10)) if config else 10
    logger.info(f"⏱️  Global Wait Time: {global_wait_time}초")
    
    results = []
    idx = 0  # 수동 인덱스 관리 (0-based)
    
    while idx < len(actions_list):
        action = actions_list[idx]
        
        # 순번이 없는 행은 스킵
        action_num = action.get('번호') or action.get('순번') or action.get('number')
        if not action_num or str(action_num).strip() == '':
            idx += 1
            continue
        
        action_num_str = str(action_num).strip()
        action_name = action.get('액션명') or action.get('name') or f'Action-{action_num_str}'
        a_type = (action.get('타입') or action.get('type') or '').lower()
        
        logger.info(f"[{action_num_str}] 실행: {action_name} (타입: {a_type})")
        
        # perform_ui_action 호출 시 actions_list와 current_index 전달
        result = perform_ui_action(
            driver, 
            action, 
            config=config, 
            test_mode=test_mode, 
            timeout=global_wait_time,
            actions_list=actions_list,
            current_index=idx
        )
        
        result['action_name'] = action_name
        result['action_index'] = action_num_str  # Excel 순번 사용
        results.append(result)

        if not test_mode:
            pause = random.uniform(1, 5)
            time.sleep(pause)
            logger.info(f"  ⏸ 랜덤 대기: {pause:.2f}초")
        
        if result['status'] == 'ok':
            logger.info(f"  ✓ {result['message']}")
        else:
            logger.error(f"  ✗ {result['message']}")
            # 에러 발생 시에도 계속 진행할지 결정 (현재는 계속 진행)
            # break  # 첫 에러에서 중단하려면 주석 해제
        
        # loop_click이 반환한 skip 정보 확인
        skip_count = result.get('skip_next_actions', 0)
        if skip_count > 0:
            logger.info(f"  ⏩ 다음 {skip_count}개 액션 건너뜀 (서브 루프에서 이미 실행됨)")
            idx += skip_count + 1  # 현재 액션(loop_click) + 서브 액션들 건너뜀
        else:
            idx += 1  # 일반 액션: 다음으로 이동
    
    logger.info(f"\n{'='*70}")
    logger.info(f"실행 완료: {len(results)}개 액션")
    success_count = sum(1 for r in results if r['status'] == 'ok')
    error_count = len(results) - success_count
    logger.info(f"  ✓ 성공: {success_count}개")
    logger.info(f"  ✗ 실패: {error_count}개")
    
    return results


def login_website(driver, url, login_id, login_pw, login_button_xpath):
    """웹사이트 접속 및 로그인"""
    
    if not driver:
        logger.error("❌ 드라이버가 초기화되지 않았습니다")
        return False
    
    logger.info(f"웹사이트 접속 중: {url}")
    
    try:
        driver.get(url)
        
        # 페이지 로드 대기 (최대 10초)
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        
        logger.info(f"✓ 접속 성공: {driver.title}")
        
        # 로그인 버튼 클릭
        logger.info("로그인 버튼 클릭...")
        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, login_button_xpath))
        )
        try:
            login_button.click()
        except ElementClickInterceptedException as e:
            def retry_login_btn_click():
                driver.execute_script("arguments[0].click();", login_button)
                logger.info("✓ JavaScript 클릭 성공")
            handle_action_error(e, 'click', login_button_xpath, retry_login_btn_click)
        time.sleep(1)
        
        # ID 입력
        logger.info(f"사용자 ID 입력: {login_id}")
        id_input = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.NAME, "userId"))
        )[0]
        id_input.clear()
        id_input.send_keys(login_id)
        time.sleep(0.5)
        
        # 비밀번호 입력
        logger.info("비밀번호 입력")
        pw_input = driver.find_element(By.NAME, "userPwd")
        pw_input.clear()
        pw_input.send_keys(login_pw)
        time.sleep(0.5)
        
        # 로그인 버튼 클릭
        logger.info("로그인 실행...")
        submit_btn = driver.find_element(By.XPATH, '//button[@type="submit"]')
        try:
            submit_btn.click()
        except ElementClickInterceptedException as e:
            def retry_submit_click():
                driver.execute_script("arguments[0].click();", submit_btn)
                logger.info("✓ JavaScript 클릭 성공")
            handle_action_error(e, 'click', '//button[@type="submit"]', retry_submit_click)
        
        # 로그인 완료 대기
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        
        logger.info(f"✓ 로그인 성공: {driver.title}")
        return True
        
    except Exception as e:
        logger.error(f"❌ 로그인 실패: {e}")
        return False


# %% 단계 3: 메인 실행

# 글로벌 변수: 재사용 가능한 드라이버 및 설정
driver = None
config_data = None
actions_data = None
TEST_MODE = False  # 실제 모드: 브라우저에서 클릭/입력 수행

# 로그인 플래그 파일 (프로세스간 상태 유지)
LOGIN_FLAG_FILE = Path(__file__).parent / ".login_flag"

def should_skip_login():
    """로그인을 건너뛸지 판단"""
    return LOGIN_FLAG_FILE.exists()

def set_login_done():
    """로그인 완료 플래그 저장"""
    LOGIN_FLAG_FILE.touch()

def reset_login_flag():
    """로그인 플래그 초기화 (다시 로그인하려면 이 함수 실행)"""
    if LOGIN_FLAG_FILE.exists():
        LOGIN_FLAG_FILE.unlink()
        logger.info("🔄 로그인 플래그 초기화됨 - 다음 실행시 다시 로그인됩니다")


# %% [Step 1] Excel 파일 읽기 및 설정 초기화
def setup_config():
    """Excel 파일을 읽어 설정을 초기화합니다. (로그인 전 한 번만 실행)"""
    global config_data, actions_data
    
    logger.info("=" * 70)
    logger.info("K-Startup 웹 자동화 - Step 1: 설정 초기화")
    logger.info("=" * 70)
    
    # 필수 라이브러리 확인
    if not OPENPYXL_AVAILABLE or not SELENIUM_AVAILABLE:
        logger.error("필수 라이브러리가 없습니다")
        return False
    
    # Excel 파일 읽기
    logger.info("[Step 1] Excel 파일 읽기")
    logger.info("-" * 70)
    
    # excel_path = Path(__file__).parent / "창업활동비.xlsx"
    excel_path = Path(__file__).parent / "회계감사비.xlsx"
    config_data, actions_data = read_excel_file(str(excel_path))
    
    if not config_data or not actions_data:
        logger.error("❌ Excel 데이터를 읽을 수 없습니다")
        return False
    
    # 설정 정보 출력
    target_url = config_data.get("base_url", "https://www.k-startup.go.kr")
    user_id = config_data.get("login_id", "test_user")
    
    logger.info(f"\n📋 추출된 설정 정보:")
    logger.info(f"  - 대상 사이트: {target_url}")
    logger.info(f"  - 사용자 ID: {user_id}")
    logger.info(f"  - 액션 개수: {len(actions_data)}")
    
    # 📂 업로드 폴더 파일 검증
    logger.info(f"\n📂 업로드 폴더 파일 검증:")
    logger.info("-" * 70)
    
    # 파일 목록 수집 (setup_config에서 한 번만 함)
    collected_files = collect_files_from_config(config_data)
    
    # 수집된 파일을 config_data에 저장 (loop_click에서 재사용)
    config_data['_collected_files'] = collected_files
    
    logger.info("-" * 70)
    logger.info("✓ 설정 초기화 완료\n")
    
    return True

# %% [실행] Step 1 실행
# ⚠️ 처음 1회: 파일 맨 위의 첫 번째 셀(임포트 부분)을 먼저 실행하세요!
# 그 다음 아래 주석을 해제하고 실행
setup_config()


# %% [Step 2] 로그인 (세션 유지)
def login():
    """로그인을 수행하고 세션을 유지합니다. (로그인 한 번만 실행)"""
    global driver, config_data
    
    if config_data is None:
        logger.error("❌ 먼저 [Step 1] 설정 초기화를 실행하세요")
        return False
    
    logger.info("=" * 70)
    logger.info("K-Startup 웹 자동화 - Step 2: 로그인")
    logger.info("=" * 70)
    
    # 브라우저 초기화
    logger.info("[Step 2-1] 브라우저 초기화")
    logger.info("-" * 70)
    
    driver = init_browser(headless=False)
    
    if not driver:
        logger.error("❌ 브라우저를 시작할 수 없습니다")
        return False
    
    # Excel 액션 1-7번으로 로그인 처리 (페이지이동, 로그인버튼, ID입력, PW입력, 로그인실행 등)
    logger.info("[Step 2-2] Excel 액션으로 로그인 (1-7번)")
    logger.info("-" * 70)
    logger.info("💡 팁: 팝업이 안 나타나면 Excel에서 액션 12번의 번호를 비우세요 (건너뜀)")
    
    # 로그인 액션 실행 전 현재 윈도우 핸들 기록
    try:
        handle_before = driver.current_window_handle
        logger.info(f"창 핸들(로그인 전): {handle_before}")
    except Exception as e:
        handle_before = None
        logger.warning(f"⚠️ 창 핸들 조회 실패(로그인 전): {e}")

    # 액션 1-7번만 실행 (로그인 과정)
    login_actions = actions_data[0:7]  # 1번부터 7번까지
    run_actions(driver, login_actions, config=config_data, test_mode=TEST_MODE)
    
    # 로그인 후 새 창/탭이 열렸다면 최신 창으로 전환
    switch_to_latest_window(driver)

    # 로그인 후 창 핸들 변경 여부 확인
    try:
        handle_after = driver.current_window_handle
        logger.info(f"창 핸들(로그인 후): {handle_after}")
        if handle_before and handle_after and handle_before != handle_after:
            logger.info(f"✓ 창 핸들 변경 감지: {handle_before} → {handle_after}")
    except Exception as e:
        logger.warning(f"⚠️ 창 핸들 조회 실패(로그인 후): {e}")

    # 로그인 상태 확인 (성공 여부 로깅)
    is_logged_in(driver)

    time.sleep(2)
    logger.info(f"✓ 로그인 완료: {driver.title}")
    
    # 로그인 후 페이지 상세 정보 출력 (요소 찾기 문제 진단용)
    get_page_info(driver)
    logger.info("\n💡 다음 단계에서 요소를 찾을 수 없으면:")
    logger.info("   1. 위의 '페이지 상세 정보'를 확인하세요")
    logger.info("   2. 개발자 도구(F12)를 열어 XPath/CSS를 다시 확인하세요")
    logger.info("   3. 필요시 Excel 설정의 XPath/CSS를 수정하세요\n")
    
    # "오늘 하루 그만 보기" 팝업 닫기
    time.sleep(1)
    try:
        logger.info("팝업 닫기 시도...")
        close_btn = driver.find_element(By.XPATH, "//*[contains(text(), '오늘 하루 그만 보기')]")
        try:
            close_btn.click()
            logger.info("✓ 팝업 닫음")
        except ElementClickInterceptedException as e:
            def retry_popup_close():
                driver.execute_script("arguments[0].click();", close_btn)
                logger.info("✓ 팝업 닫음 (JavaScript 사용)")
            handle_action_error(e, 'click', "팝업 닫기 버튼", retry_popup_close)
    except NoSuchElementException:
        logger.info("⚠️ 팝업 없음 (Excel에서 액션 12번 번호를 비우면 건너뜀)")
    except Exception as ex:
        logger.warning(f"⚠️ 팝업 처리 중 예외: {ex}")
    
    logger.info("\n✓ 로그인 완료 (세션 유지 중)\n")
    return True


# %% [실행] Step 2 실행
# 테스트 모드에서는 실제 로그인/브라우저 동작을 건너뜁니다
if not TEST_MODE:
    login()  # ← 실 실행 모드에서만 로그인 수행
else:
    logger.info("🧪 TEST MODE - 로그인 단계 건너뜀 (브라우저 미시작)")


# %% [Step 3] 액션 실행 (로그인 후)
def execute_actions():
    """로그인 후 액션을 실행합니다. (반복 실행 가능)"""
    global driver, config_data, actions_data, TEST_MODE
    
    if driver is None and not TEST_MODE:
        logger.error("❌ 먼저 [Step 2] 로그인을 실행하세요")
        return
    
    if config_data is None or actions_data is None:
        logger.error("❌ 먼저 [Step 1] 설정 초기화를 실행하세요")
        return
    
    logger.info("=" * 70)
    logger.info("K-Startup 웹 자동화 - Step 3: 액션 실행 (8번부터)")
    if TEST_MODE:
        logger.info("🧪 TEST MODE 활성화 (로그만 출력)")
    logger.info("=" * 70)
    
    # 액션 8번부터 실행 (로그인은 Step 2에서 완료)
    logger.info("[Step 3] 액션 8번부터 실행")
    logger.info("-" * 70)
    
    # 실행 전 페이지 상태 출력 (TEST MODE는 건너뜀)
    if not TEST_MODE:
        logger.info("\n📊 현재 페이지 상태:")
        get_page_info(driver)
    
    test_actions = actions_data[7:]  # 8번부터 끝까지
    # 로그인 후 다른 창/탭으로 전환되었다면 최신 창을 활성화 (TEST MODE는 건너뜀)
    if not TEST_MODE:
        switch_to_latest_window(driver)

    # 로그인 상태 점검: 이미 로그인 상태면 액션 8번부터 진행, 아니면 경고
    if not TEST_MODE:
        if not is_logged_in(driver):
            logger.warning("⚠️ 현재 로그인 상태를 확인하지 못했습니다. Step 2(로그인)을 다시 실행하거나 수동 로그인 후 재시도하세요.")

    results = run_actions(driver, test_actions, config=config_data, test_mode=TEST_MODE)
    
    logger.info("\n" + "=" * 70)
    logger.info("Step 3 완료")
    logger.info("=" * 70)


# %% [실행] Step 3 실행 (반복 가능!)
execute_actions()  # ← 이 주석을 해제하고 셀 실행 (여러 번 가능)


# %% [Step 4] 종료 (옵션)
def cleanup():
    """브라우저를 종료합니다."""
    global driver
    
    if driver:
        logger.info("=" * 70)
        logger.info("K-Startup 웹 자동화 - Step 4: 종료")
        logger.info("=" * 70)
        
        try:
            driver.quit()
            driver = None
            logger.info("✓ 브라우저 종료")
        except Exception as e:
            logger.error(f"❌ 브라우저 종료 실패: {e}")
    else:
        logger.info("⚠️  활성 브라우저가 없습니다")


# %% [실행] Step 4 실행
cleanup()  # ← 이 주석을 해제하고 셀 실행


# %% 헬퍼: 전체 자동 실행 함수
def main():
    """자동 실행 모드: 로그인은 한 번만, 액션은 반복 실행"""
    global driver
    
    if setup_config():
        # TEST MODE: 브라우저 없이 전체 시뮬레이션
        if TEST_MODE:
            logger.info("🧪 TEST MODE - 브라우저 없이 액션 시뮬레이션 실행")
            execute_actions()
            logger.info("\n✓ 테스트 시뮬레이션 완료")
            logger.info("💡 실제 실행은 TEST_MODE를 False로 바꾸세요")
            return

        # 로그인 건너뛰기 확인 (파일 기반)
        skip_login = should_skip_login()
        
        if not skip_login:
            # 로그인 필요
            if login():
                set_login_done()  # 로그인 완료 플래그 저장
                logger.info("✓ 로그인 완료 - 다음 실행부터는 로그인 건너뜀")
            else:
                logger.error("❌ 로그인 실패")
                return
        else:
            if driver is None:
                logger.warning("⚠️ 드라이버가 없습니다. 로그인 플래그를 초기화하고 다시 시도하세요.")
                logger.info("   팁: Python 콘솔에서 reset_login_flag() 실행")
                return
            logger.info("⏭️ 로그인 건너뜀 (이미 로그인됨)")
        
        execute_actions()
        # cleanup()  # 테스트 후 브라우저 종료 원할 때 주석 해제
        
        logger.info("\n✓ 프로그램 완료")
        logger.info("💡 다시 로그인하려면 콘솔에서: reset_login_flag() 실행")


# %% 프로그램 시작점 (자동 실행)
if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logger.info("\n프로그램이 사용자에 의해 중단되었습니다")
        cleanup()
    except Exception as e:
        logger.error(f"예상치 못한 오류: {e}", exc_info=True)
        cleanup()


# %% [디버그] 요소 찾기 문제 진단 유틸리티
# 요소를 찾을 수 없을 때 실행하여 현재 페이지 상태를 확인하세요

def show_page_status():
    """현재 페이지의 상세 정보를 출력합니다."""
    if driver is None:
        logger.error("❌ 먼저 로그인을 실행하세요")
        return
    
    logger.info("\n" + "=" * 70)
    logger.info("📊 현재 페이지 상태 분석")
    logger.info("=" * 70)
    get_page_info(driver)


def show_all_links():
    """현재 페이지의 모든 링크를 나열합니다."""
    if driver is None:
        logger.error("❌ 먼저 로그인을 실행하세요")
        return
    
    logger.info("\n" + "=" * 70)
    logger.info("🔗 현재 페이지의 모든 링크")
    logger.info("=" * 70)
    list_all_elements(driver, tag_name='a')


def show_all_buttons():
    """현재 페이지의 모든 버튼을 나열합니다."""
    if driver is None:
        logger.error("❌ 먼저 로그인을 실행하세요")
        return
    
    logger.info("\n" + "=" * 70)
    logger.info("🔘 현재 페이지의 모든 버튼")
    logger.info("=" * 70)
    list_all_elements(driver, tag_name='button')


def diagnose_element(xpath):
    """특정 요소의 찾기 여부를 진단합니다."""
    if driver is None:
        logger.error("❌ 먼저 로그인을 실행하세요")
        return
    
    logger.info("\n" + "=" * 70)
    logger.info(f"🔍 XPath 진단: {xpath}")
    logger.info("=" * 70)
    diagnose_element_issue(driver, xpath)


# 사용 예시:
# show_page_status()  # 페이지 상태 확인
# show_all_links()    # 모든 링크 보기
# show_all_buttons()  # 모든 버튼 보기
# diagnose_element("//button[contains(text(), '저장')]")  # 특정 요소 진단
