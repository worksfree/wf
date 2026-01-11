# %% [셀 0] 라이브러리 임포트
# -*- coding: utf-8 -*-
"""
K-Startup 웹 자동화 (Interactive ver.)
"""
import sys
import time
import logging
from pathlib import Path
from datetime import datetime

# UTF-8 출력 (Windows CMD 호환)
try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
except: pass

# 필수 라이브러리
try:
    from openpyxl import load_workbook
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        TimeoutException, 
        ElementClickInterceptedException
    )
    import pyperclip
    import pyautogui
except ImportError as e:
    print(f"❌ 필수 라이브러리 부족: {e}")
    print("pip install selenium openpyxl pyperclip pyautogui")

# 로거 설정
logger = logging.getLogger("kstartup_interactive")
logger.setLevel(logging.INFO)
if not logger.handlers:
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    fmt = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', '%H:%M:%S')
    ch.setFormatter(fmt)
    logger.addHandler(ch)

logger.info("✅ 라이브러리 로드 완료")


# %% [셀 1] Excel 설정 로드 (회계감사비.xlsx)
config_data = {}
actions_data = []

def load_excel():
    global config_data, actions_data
    # excel_file = Path(__file__).parent / "회계감사비.xlsx"
    excel_file = Path(r"D:\drive_files\10.worksfree\10.rpa\70.webs\kstartup-web\회계감사비.xlsx")
    
    if not excel_file.exists():
        logger.error(f"❌ Excel 파일 없음: {excel_file}")
        return

    wb = load_workbook(excel_file, data_only=True)
    
    # 1. Config 시트
    if 'Config' in wb.sheetnames:
        ws = wb['Config']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]: config_data[row[0]] = row[1]
    
    # 2. Actions 시트
    if 'Actions' in wb.sheetnames:
        ws = wb['Actions']
        headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        actions_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            action = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
            actions_data.append(action)

    logger.info(f"✅ Excel 로드 완료")
    logger.info(f"   - URL: {config_data.get('base_url')}")
    logger.info(f"   - ID: {config_data.get('login_id')}")
    logger.info(f"   - Actions: {len(actions_data)}개")

load_excel()


# %% [셀 2] 브라우저 시작 및 로그인 (1회만 실행!)
driver = None

def start_and_login():
    global driver
    if driver:
        logger.warning("⚠️ 이미 브라우저가 실행 중입니다.")
        return

    # 옵션 설정 (Headless 아님!)
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized") 
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("detach", True) # 프로세스 종료 방지
    
    # 여기서 headless 절대 사용 안함
    # options.add_argument("--headless") 

    try:
        driver = webdriver.Chrome(options=options)
        logger.info("✅ 브라우저 시작됨")
    except Exception as e:
        logger.error(f"❌ 브라우저 시작 에러: {e}")
        return

    # 로그인 로직
    try:
        url = config_data.get('base_url', 'https://www.k-startup.go.kr/')
        uid = config_data.get('login_id', '')
        upw = config_data.get('login_pw', '')

        if not uid or not upw:
            logger.error("❌ ID/PW가 설정되지 않음")
            return

        driver.get(url)
        logger.info(f"🚀 사이트 접속: {url}")
        
        # 로그인 폼 대기
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "userEmail"))
        )
        
        # 입력
        driver.find_element(By.NAME, "userEmail").send_keys(uid)
        time.sleep(0.5)
        driver.find_element(By.NAME, "userPwd").send_keys(upw)
        time.sleep(0.5)
        
        # 로그인 버튼 클릭 시도 (3단계)
        login_btn_xpath = '//button[@type="submit"]'
        
        # 1. 일반 클릭
        try:
            btn = driver.find_element(By.XPATH, login_btn_xpath)
            btn.click()
            logger.info("👉 [1] 일반 클릭 성공")
        except:
            # 2. JS 클릭
            try:
                btn = driver.find_element(By.XPATH, login_btn_xpath)
                driver.execute_script("arguments[0].click();", btn)
                logger.info("👉 [2] JS 클릭 성공")
            except:
                # 3. 엔터키
                driver.find_element(By.NAME, "userPwd").send_keys(Keys.RETURN)
                logger.info("👉 [3] 엔터키 입력")
        
        time.sleep(3)
        logger.info(f"✅ 로그인 단계 완료. 현재 제목: {driver.title}")

    except Exception as e:
        logger.error(f"❌ 로그인 중 오류: {e}")

start_and_login()


# %% [셀 3] 액션 실행 (반복 가능)
def run_actions(start_idx=7):
    if not driver:
        logger.error("❌ 브라우저가 없습니다. 셀 2를 실행하세요.")
        return

    logger.info(f"🚀 액션 실행 시작 (인덱스 {start_idx}부터)")
    
    # 로그인(0~6) 이후부터 실행
    targets = actions_data[start_idx:]
    
    for i, act in enumerate(targets):
        idx = start_idx + i + 1
        name = act.get('액션명', '알수없음')
        atype = act.get('타입', '').lower()
        xpath = act.get('엘리먼트')
        val = act.get('값')
        
        logger.info(f"[{idx}] {name} ({atype})")
        
        try:
            if atype == 'click':
                el = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                try:
                    el.click()
                except:
                    driver.execute_script("arguments[0].click();", el)
                logger.info("   ✓ 클릭 완료")
                
            elif atype == 'input':
                el = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                el.clear()
                el.send_keys(str(val))
                logger.info(f"   ✓ 입력 완료: {val}")
                
            elif atype == 'sleep':
                sec = float(val) if val else 1.0
                time.sleep(sec)
                logger.info(f"   💤 대기 {sec}초")
                
            time.sleep(1)
            
        except Exception as e:
            logger.error(f"   ❌ 실패: {e}")

# 실행 (7번 인덱스 = 8번째 액션부터)
run_actions(7)
